"""
检测报告导出模块 - 范本风格三层结构
格式：宋体，等宽列14.0×8列，thin border全覆盖
"""
from pathlib import Path
from datetime import datetime
import os
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 全局样式常量
# ============================================================
THIN = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FILL = PatternFill('solid', fgColor='D9E1F2')  # 浅蓝底色
COL_COUNT = 8
COL_WIDTH = 14.0
SIGN_ROW_HEIGHT = 35


def _font(bold=False, size=11):
    """宋体字体"""
    return Font(name='宋体', size=size, bold=bold)


def _cell(ws, row, col, value='', *, bold=False, size=10, align='center', fill=None):
    """写入单元格并设置样式"""
    c = ws.cell(row, col, value)
    c.font = _font(bold, size)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = BORDER
    if fill:
        c.fill = fill
    return c


def _merge(ws, r1, c1, r2, c2, value='', *, bold=False, size=10, align='center', fill=None):
    """合并单元格并写入值"""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    _cell(ws, r1, c1, value, bold=bold, size=size, align=align, fill=fill)
    # 给合并区域所有单元格加边框
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if fill:
                cell.fill = fill


def _set_base_layout(ws):
    """设置基础布局：8列等宽14.0，隐藏网格线"""
    for i in range(1, COL_COUNT + 1):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTH
    ws.sheet_view.showGridLines = False


# ============================================================
# 数据提取函数（保持不变）
# ============================================================
def _clean_text(v):
    return str(v or '').replace('\u2705', '').replace('\u274c', '').replace('\u26a0\ufe0f', '').replace('\u26a0', '').strip()


def _judge_from_result(v):
    txt = str(v or '').strip()
    if not txt:
        return ''
    if any(flag in txt for flag in ('❌', '✗', '☒')):
        return '不合格'
    if any(flag in txt for flag in ('✅', '✓', '☑')):
        return '合格'
    normalized = txt.replace('（', '(').replace('）', ')').replace(' ', '')
    fail_markers = ['不合格', '不符合要求', '未通过', '超标', '异常']
    pass_markers = ['合格', '符合要求', '通过']
    if any(marker in normalized for marker in fail_markers):
        return '不合格'
    if any(marker in normalized for marker in pass_markers):
        return '合格'
    return ''


def _judge_for_key(result_text, *param_keys):
    """优先使用 judgement_result 索引（后端引擎权威判定），
    找不到时 fallback 到 result 文本解析（emoji/关键词）。
    """
    # 优先：judgement_result 索引
    if _current_judgement_index and param_keys:
        for k in param_keys:
            if k in _current_judgement_index:
                return '合格' if _current_judgement_index[k] else '不合格'
            # 前缀匹配
            matched_any = False
            all_passed = True
            for idx_key, idx_passed in _current_judgement_index.items():
                if idx_key == k or idx_key.startswith(k + '.'):
                    matched_any = True
                    if not idx_passed:
                        all_passed = False
                        break
            if matched_any:
                return '合格' if all_passed else '不合格'
    # fallback：从 result 文本解析
    judged = _judge_from_result(result_text)
    if judged:
        return judged
    return ''


def _normalize_param_item(item):
    if isinstance(item, dict):
        if 'value' in item and 'values' not in item:
            value = item.get('value')
            item = {**item, 'values': [] if value in (None, '') else [value]}
        if 'result' not in item:
            vals = item.get('values')
            if isinstance(vals, list) and vals:
                item = {**item, 'result': vals[0]}
        return item
    if item in (None, ''):
        return {}
    return {'values': [item], 'result': item}


def _get_param(room, *keys):
    params = room.get('params', {}) or {}
    alias_map = {
        'settling': ('settling', 'settle_bacteria', 'bacteria'),
        'floating': ('floating', 'floating_bacteria'),
        'self_purification_time': ('self_purification_time', 'self_purify_time', 'self_clean'),
        'illumination': ('illumination', 'illumination_main_room', 'illumination_main', 'illumination_mixed_processing'),
    }
    search_keys = []
    for key in keys:
        aliases = alias_map.get(key, (key,))
        for a in aliases:
            if a not in search_keys:
                search_keys.append(a)

    if isinstance(params, dict):
        for key in search_keys:
            if key in params:
                return _normalize_param_item(params[key])
    elif isinstance(params, list):
        for item in params:
            if not isinstance(item, dict):
                continue
            if item.get('key') in search_keys:
                return _normalize_param_item(item)
    return {}


def _result(room, *keys, default=''):
    p = _get_param(room, *keys)
    raw = _clean_text(p.get('result', default))
    if raw and not _judge_from_result(raw):
        values = p.get('values')
        if isinstance(values, list) and values:
            joined = ','.join(str(v) for v in values if str(v).strip())
            if joined and raw in joined:
                raw = ''
    return raw


def _build_judgement_index(export_payload):
    """从 judgement_result.item_results 建立 {key: passed} 索引。
    支持两种结构：
      - item_results 是 list[dict]，每个 dict 有 key + passed
      - item_results 是 dict，key → {passed: bool}
    """
    jr = export_payload.get('judgement_result') or {}
    items = jr.get('item_results') or []
    index = {}
    if isinstance(items, list):
        for item in items:
            if isinstance(item, dict) and 'key' in item and 'passed' in item:
                index[item['key']] = item['passed']
    elif isinstance(items, dict):
        for k, v in items.items():
            if isinstance(v, dict) and 'passed' in v:
                index[k] = v['passed']
            elif isinstance(v, bool):
                index[k] = v
    return index


# 模块级缓存，由各 _build_xxx_record 函数在开头设置
_current_judgement_index = {}


def _judge_text(room, *keys, default=''):
    """判定文本：优先从 judgement_result 索引取，再 fallback 到 result 文本解析。"""
    # 优先查 judgement_result 索引
    if _current_judgement_index and keys:
        # 尝试多种 key 格式匹配
        for k in keys:
            if k in _current_judgement_index:
                return '合格' if _current_judgement_index[k] else '不合格'
            # 尝试带前缀的 key，如 particle.op_05_max
            for idx_key, passed in _current_judgement_index.items():
                if idx_key.startswith(k + '.') or idx_key == k:
                    # 只要有一个不合格就不合格
                    if not passed:
                        return '不合格'

    p = _get_param(room, *keys)
    # 先用原始 result（含 emoji）做判定，再用 clean 版本做 fallback
    raw_result = p.get('result', default)
    judged = _judge_from_result(raw_result)
    if judged:
        return judged
    # 再试 clean 版本（去掉 emoji 后可能暴露中文判定词）
    result_text = _clean_text(raw_result)
    judged = _judge_from_result(result_text)
    if judged:
        return judged
    # fallback: passed 布尔 (少数 param 如 airtightness 有 pass 字段)
    passed = p.get('passed') if p.get('passed') is not None else p.get('pass')
    if passed is True:
        return '合格'
    if passed is False:
        return '不合格'
    # 最后再查一次索引（用第一个 key 做前缀匹配）
    if _current_judgement_index and keys:
        main_key = keys[0]
        matched_any = False
        all_passed = True
        for idx_key, idx_passed in _current_judgement_index.items():
            if idx_key == main_key or idx_key.startswith(main_key + '.'):
                matched_any = True
                if not idx_passed:
                    all_passed = False
                    break
        if matched_any:
            return '合格' if all_passed else '不合格'
    return ''


def _values(room, *keys):
    p = _get_param(room, *keys)
    vals = p.get('values')
    if isinstance(vals, list):
        return [str(v) for v in vals if str(v).strip()]
    if vals not in (None, ''):
        return [str(vals)]
    return []


def _data(room, *keys):
    p = _get_param(room, *keys)
    return p.get('data') or {}


def _pairs(room, *keys):
    p = _get_param(room, *keys)
    return p.get('pairs') or []


def _vents(room, *keys):
    p = _get_param(room, *keys)
    return p.get('vents') or []

# ============================================================
# 判定范围提取
# ============================================================
def _get_judge_range(export_payload, param_key):
    """从 judgement_result 或 template_rule 中获取判定范围文本"""
    # 尝试从 judgement_result 获取
    jr = export_payload.get('judgement_result') or {}
    if isinstance(jr, dict):
        item = jr.get(param_key) or {}
        if isinstance(item, dict):
            rng = item.get('range') or item.get('standard') or item.get('rule_text') or ''
            if rng:
                return str(rng)
        elif isinstance(item, str) and item:
            return item

    # 尝试从 template_rule 获取
    tr = export_payload.get('template_rule') or {}
    if isinstance(tr, dict):
        rule = tr.get(param_key) or {}
        if isinstance(rule, dict):
            rng = rule.get('range') or rule.get('standard') or rule.get('text') or ''
            if rng:
                return str(rng)
        elif isinstance(rule, str) and rule:
            return rule

    return ''


# ============================================================
# 页头 & 项目信息写入
# ============================================================
def _write_page_header(ws, project, room, export_payload):
    """
    写入页头标题和项目信息区域
    返回下一个可用行号
    """
    # 第1行：页头标题 - 宋体16号加粗居中，合并A1:H1
    _merge(ws, 1, 1, 1, 8, '检测原始记录', bold=True, size=16, align='center')

    # 空行
    row = 3

    # 项目信息 - 标签字段加粗11号左对齐，值字段10号左对齐
    info_items = [
        ('项目名称', project.get('project_name', ''), '委托单位', project.get('client_name', '')),
        ('报告编号', project.get('report_number', ''), '检测日期', project.get('detection_date', '')),
        ('检测区域', room.get('room_name', '') or project.get('inspection_area', ''), '检测状态', project.get('detection_state', '') or project.get('detection_state_text', '')),
        ('检测类型', room.get('type_name', ''), '洁净等级', room.get('clean_class', '') or room.get('level_name', '')),
    ]

    for label1, val1, label2, val2 in info_items:
        _cell(ws, row, 1, f'{label1}：', bold=True, size=11, align='left')
        _merge(ws, row, 2, row, 4, val1, size=10, align='left')
        _cell(ws, row, 5, f'{label2}：', bold=True, size=11, align='left')
        _merge(ws, row, 6, row, 8, val2, size=10, align='left')
        row += 1

    # 引用标准
    basis_list = room.get('basis', []) or project.get('basis', []) or []
    basis_text = '、'.join(basis_list) if isinstance(basis_list, list) else str(basis_list)
    _cell(ws, row, 1, '检测依据：', bold=True, size=11, align='left')
    _merge(ws, row, 2, row, 8, basis_text, size=10, align='left')
    row += 1

    judge_list = room.get('judgement', []) or project.get('judgement', []) or []
    judge_text = '、'.join(judge_list) if isinstance(judge_list, list) else str(judge_list)
    _cell(ws, row, 1, '判定标准：', bold=True, size=11, align='left')
    _merge(ws, row, 2, row, 8, judge_text, size=10, align='left')
    row += 1

    return row + 1  # 空一行后开始检测参数


# ============================================================
# 三层结构检测项写入
# ============================================================
def _write_section(ws, row, title, headers, data_rows, result_text, judge_text):
    """
    写入一个检测项的三层结构：
    - 标题行：合并A:H，加粗11号，浅蓝底色
    - 测点表头行：9号字居中
    - 数据行：10号字居中
    - 结果行：检测结果(B:D) | 判定(F:H)
    返回下一个可用行号
    """
    # 标题行
    _merge(ws, row, 1, row, 8, title, bold=True, size=11, align='left', fill=TITLE_FILL)
    row += 1

    # 测点表头行 - 9号字居中
    if headers:
        for i, h in enumerate(headers[:COL_COUNT], start=1):
            _cell(ws, row, i, h, size=9, align='center')
        # 填充剩余列
        for i in range(len(headers) + 1, COL_COUNT + 1):
            _cell(ws, row, i, '', size=9)
        row += 1

    # 数据行 - 10号字居中
    for data_row in data_rows:
        for i, v in enumerate(data_row[:COL_COUNT], start=1):
            _cell(ws, row, i, str(v) if v else '', size=10, align='center')
        for i in range(len(data_row) + 1, COL_COUNT + 1):
            _cell(ws, row, i, '', size=10)
        row += 1

    # 结果行：检测结果：(加粗) + 值(B:D) | 判定：(加粗) + 合格/不合格(F:H)
    _cell(ws, row, 1, '检测结果：', bold=True, size=10, align='left')
    _merge(ws, row, 2, row, 4, result_text, size=10, align='left')
    _cell(ws, row, 5, '判定：', bold=True, size=10, align='left')
    _merge(ws, row, 6, row, 8, judge_text, size=10, align='left')
    row += 1

    return row


def _write_section_simple(ws, row, title, values_list, result_text, judge_text):
    """简化版：单行数据的检测项"""
    headers = [f'测点{i+1}' for i in range(len(values_list))] if values_list else ['测点1']
    data_rows = [values_list] if values_list else [['']]
    return _write_section(ws, row, title, headers, data_rows, result_text, judge_text)


def _write_cleanness_section(ws, row, title, zones_data, result_text, judge_text):
    """
    洁净度特殊结构：多区域 × 多粒径
    zones_data: [{'zone': '手术区', 'particles': [{'size': '>=0.5um', 'max': x, 'ucl': y}, ...]}]
    """
    # 标题行
    _merge(ws, row, 1, row, 8, title, bold=True, size=11, align='left', fill=TITLE_FILL)
    row += 1

    # 表头
    headers = ['区域', '粒径', '最大值', 'UCL', '', '', '', '']
    for i, h in enumerate(headers[:COL_COUNT], start=1):
        _cell(ws, row, i, h, size=9, align='center')
    row += 1

    # 数据行
    for zone in zones_data:
        zone_name = zone.get('zone', '')
        particles = zone.get('particles', [])
        for p in particles:
            data = [zone_name, p.get('size', ''), p.get('max', ''), p.get('ucl', ''), '', '', '', '']
            for i, v in enumerate(data[:COL_COUNT], start=1):
                _cell(ws, row, i, str(v) if v else '', size=10, align='center')
            row += 1
            zone_name = ''  # 后续行不重复区域名

    # 结果行
    _cell(ws, row, 1, '检测结果：', bold=True, size=10, align='left')
    _merge(ws, row, 2, row, 4, result_text, size=10, align='left')
    _cell(ws, row, 5, '判定：', bold=True, size=10, align='left')
    _merge(ws, row, 6, row, 8, judge_text, size=10, align='left')
    row += 1

    return row


# ============================================================
# 签名行
# ============================================================
def _write_signature(ws, row, project):
    """写入签名行，行高35"""
    _cell(ws, row, 1, '检测员：', bold=True, size=11, align='left')
    _merge(ws, row, 2, row, 4, project.get('inspector', ''), size=10, align='left')
    _cell(ws, row, 5, '校核人：', bold=True, size=11, align='left')
    _merge(ws, row, 6, row, 8, project.get('reviewer', ''), size=10, align='left')
    ws.row_dimensions[row].height = SIGN_ROW_HEIGHT
    row += 1

    _cell(ws, row, 1, '检测日期：', bold=True, size=11, align='left')
    _merge(ws, row, 2, row, 4, project.get('detection_date', ''), size=10, align='left')
    _cell(ws, row, 5, '记录日期：', bold=True, size=11, align='left')
    _merge(ws, row, 6, row, 8, datetime.now().strftime('%Y-%m-%d'), size=10, align='left')
    ws.row_dimensions[row].height = SIGN_ROW_HEIGHT
    row += 1

    return row

# ============================================================
# 医院类检测项导出
# ============================================================
def _build_hospital_record(export_payload, output_path):
    """
    医院类：negative_pressure, operating_room, clean_function_room, bsl, animal_room
    检测项：换气次数、截面平均风速、风速不均匀度、静压差、高效过滤器检漏、严密性、
           洁净度（悬浮粒子，含手术区/周边区多区域）、开门后洁净度、温度、相对湿度、
           噪声、照度（最低照度）、照度均匀度、沉降菌、浮游菌
    """
    global _current_judgement_index
    _current_judgement_index = _build_judgement_index(export_payload)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    project = export_payload.get('project', {}) or {}
    room = export_payload.get('room', {}) or {}
    ws.title = str(room.get('room_name') or '原始记录')[:31]
    _set_base_layout(ws)

    row = _write_page_header(ws, project, room, export_payload)

    # --- 换气次数 ---
    vents = _vents(room, 'airchange')
    airchange_vals = _values(room, 'airchange')
    judge_range = _get_judge_range(export_payload, 'airchange')
    title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
    if vents or airchange_vals:
        title = f'换气次数（次/h）{title_suffix}'
        if vents:
            v = vents[0]
            area = float(v.get('area') or 0)
            speed = float(v.get('speed') or 0)
            volume = float(v.get('volume') or 0)
            length = float(room.get('length') or 0)
            width = float(room.get('width') or 0)
            height = float(room.get('height') or 0)
            room_volume = round(length * width * height, 1) if length and width and height else 0
            if not volume and area and speed:
                volume = round(area * speed * 3600, 1)
            headers = ['风口面积(m²)', '风速(m/s)', '风量(m³/h)', '房间体积(m³)', '换气次数', '', '', '']
            data = [str(area), str(speed), str(volume), str(room_volume), _result(room, 'airchange'), '', '', '']
            row = _write_section(ws, row, title, headers, [data], _result(room, 'airchange'), _judge_text(room, 'airchange'))
        else:
            row = _write_section_simple(ws, row, title, airchange_vals, _result(room, 'airchange'), _judge_text(room, 'airchange'))

    # --- 截面平均风速 ---
    wind_vals = _values(room, 'wind_speed', 'air_velocity')
    if wind_vals:
        judge_range = _get_judge_range(export_payload, 'wind_speed')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'截面平均风速（m/s）{title_suffix}'
        row = _write_section_simple(ws, row, title, wind_vals, _result(room, 'wind_speed', 'air_velocity'), _judge_for_key(_result(room, 'wind_speed', 'air_velocity'), 'wind_speed'))

    # --- 风速不均匀度 ---
    uniformity_vals = _values(room, 'wind_uniformity', 'velocity_uniformity')
    if uniformity_vals:
        judge_range = _get_judge_range(export_payload, 'wind_uniformity')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'风速不均匀度{title_suffix}'
        row = _write_section_simple(ws, row, title, uniformity_vals, _result(room, 'wind_uniformity', 'velocity_uniformity'), _judge_for_key(_result(room, 'wind_uniformity', 'velocity_uniformity'), 'wind_uniformity'))

    # --- 排风口风速（负压病房） ---
    exhaust_vals = _values(room, 'exhaust_speed')
    exhaust_result = _result(room, 'exhaust_speed')
    if exhaust_vals or exhaust_result:
        judge_range = _get_judge_range(export_payload, 'exhaust_speed')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'排风口风速（m/s）{title_suffix}'
        row = _write_section_simple(ws, row, title, exhaust_vals or [exhaust_result], exhaust_result, _judge_for_key(exhaust_result, 'exhaust_speed'))

    # --- 笼具工作面风速（动物房） ---
    cage_vals = _values(room, 'cage_airspeed')
    cage_result = _result(room, 'cage_airspeed')
    if cage_vals or cage_result:
        judge_range = _get_judge_range(export_payload, 'cage_airspeed')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'笼具工作面风速（m/s）{title_suffix}'
        row = _write_section_simple(ws, row, title, cage_vals or [cage_result], cage_result, _judge_for_key(cage_result, 'cage_airspeed'))

    # --- 静压差 ---
    prs = _pairs(room, 'pressure')
    pressure_vals = _values(room, 'pressure')
    if prs or pressure_vals:
        judge_range = _get_judge_range(export_payload, 'pressure')
        # 判断正压/负压
        type_id = str(room.get('type_id', '') or '')
        pressure_type = 'negative' if type_id == 'negative_pressure' else 'positive'
        pressure_label = '负压' if pressure_type == 'negative' else '正压'
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'静压差（{pressure_label}）（Pa）{title_suffix}'
        if prs:
            # 多对压差
            headers = ['相对区域', '测点1', '测点2', '测点3', '平均值', '', '', '']
            data_rows = []
            for pair in prs:
                ref = pair.get('ref', '') or pair.get('reference', '')
                vals = pair.get('values') or []
                avg = pair.get('avg', '') or pair.get('result', '')
                row_data = [ref] + [str(v) for v in vals[:3]] + [str(avg)] + [''] * (8 - 5)
                data_rows.append(row_data[:8])
            row = _write_section(ws, row, title, headers, data_rows, _result(room, 'pressure'), _judge_for_key(_result(room, 'pressure'), 'pressure'))
        else:
            row = _write_section_simple(ws, row, title, pressure_vals, _result(room, 'pressure'), _judge_for_key(_result(room, 'pressure'), 'pressure'))

    # --- 高效过滤器检漏 ---
    hepa = _get_param(room, 'hepa_leak')
    if hepa:
        judge_range = _get_judge_range(export_payload, 'hepa_leak')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'高效过滤器检漏（%）{title_suffix}'
        hepa_vals = _values(room, 'hepa_leak')
        if not hepa_vals:
            total = _data(room, 'hepa_leak').get('total', '')
            hepa_vals = [str(total)] if total not in (None, '') else ['']
        row = _write_section_simple(ws, row, title, hepa_vals, _result(room, 'hepa_leak'), _judge_for_key(_result(room, 'hepa_leak'), 'hepa_leak'))

    # --- 严密性 ---
    airtight_vals = _values(room, 'airtightness', 'tightness')
    if airtight_vals or _result(room, 'airtightness', 'tightness'):
        judge_range = _get_judge_range(export_payload, 'airtightness')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'严密性{title_suffix}'
        result = _result(room, 'airtightness', 'tightness') or '符合要求'
        row = _write_section_simple(ws, row, title, airtight_vals or [result], result, _judge_for_key(result, 'airtightness'))

    # --- 洁净度（悬浮粒子） ---
    pdata = _data(room, 'particle')
    particle_vals = _values(room, 'particle')
    particle_param = _get_param(room, 'particle')
    if pdata or particle_vals or particle_param:
        judge_range = _get_judge_range(export_payload, 'particle')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'洁净度（悬浮粒子）（粒/m³）{title_suffix}'

        # 检查是否有多区域数据（手术区/周边区）
        zones = particle_param.get('zones') or pdata.get('zones') or []
        if zones:
            zones_data = []
            for z in zones:
                zone_name = z.get('zone', '') or z.get('name', '')
                particles = []
                for size_key in ['p05', 'p5', '0.5', '5']:
                    max_val = z.get(f'{size_key}_max', '') or z.get(f'max_{size_key}', '')
                    ucl_val = z.get(f'{size_key}_ucl', '') or z.get(f'ucl_{size_key}', '')
                    if max_val or ucl_val:
                        size_label = '≥0.5μm' if '05' in size_key or '0.5' in size_key else '≥5μm'
                        particles.append({'size': size_label, 'max': max_val, 'ucl': ucl_val})
                # 也检查 particles 列表格式
                if not particles and 'particles' in z:
                    for p in z['particles']:
                        particles.append({'size': p.get('size', ''), 'max': p.get('max', ''), 'ucl': p.get('ucl', '')})
                if particles:
                    zones_data.append({'zone': zone_name, 'particles': particles})

            if zones_data:
                row = _write_cleanness_section(ws, row, title, zones_data, _result(room, 'particle'), _judge_for_key(particle_param.get('result', ''), 'particle'))
            else:
                row = _write_section_simple(ws, row, title, particle_vals, _result(room, 'particle'), _judge_for_key(particle_param.get('result', ''), 'particle'))
        elif pdata:
            # 标准粒子数据格式
            zones_data = [{'zone': '检测区域', 'particles': []}]
            if pdata.get('p05_max') or pdata.get('p05_ucl'):
                zones_data[0]['particles'].append({'size': '≥0.5μm', 'max': pdata.get('p05_max', ''), 'ucl': pdata.get('p05_ucl', '')})
            if pdata.get('p5_max') or pdata.get('p5_ucl'):
                zones_data[0]['particles'].append({'size': '≥5μm', 'max': pdata.get('p5_max', ''), 'ucl': pdata.get('p5_ucl', '')})
            if zones_data[0]['particles']:
                row = _write_cleanness_section(ws, row, title, zones_data, _result(room, 'particle'), _judge_for_key(particle_param.get('result', ''), 'particle'))
            else:
                row = _write_section_simple(ws, row, title, particle_vals, _result(room, 'particle'), _judge_for_key(particle_param.get('result', ''), 'particle'))
        else:
            row = _write_section_simple(ws, row, title, particle_vals, _result(room, 'particle'), _judge_for_key(particle_param.get('result', ''), 'particle'))

    # --- 开门后洁净度 ---
    door_open_vals = _values(room, 'particle_door_open', 'door_open_particle')
    if door_open_vals or _result(room, 'particle_door_open', 'door_open_particle'):
        judge_range = _get_judge_range(export_payload, 'particle_door_open')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'开门后洁净度（粒/m³）{title_suffix}'
        row = _write_section_simple(ws, row, title, door_open_vals or [''], _result(room, 'particle_door_open', 'door_open_particle'), _judge_for_key(_result(room, 'particle_door_open', 'door_open_particle'), 'particle_door_open'))

    # --- 温度 ---
    temp_vals = _values(room, 'temperature')
    if temp_vals:
        judge_range = _get_judge_range(export_payload, 'temperature')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'温度（℃）{title_suffix}'
        row = _write_section_simple(ws, row, title, temp_vals, _result(room, 'temperature'), _judge_text(room, 'temperature'))

    # --- 相对湿度 ---
    humid_vals = _values(room, 'humidity')
    if humid_vals:
        judge_range = _get_judge_range(export_payload, 'humidity')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'相对湿度（%）{title_suffix}'
        row = _write_section_simple(ws, row, title, humid_vals, _result(room, 'humidity'), _judge_text(room, 'humidity'))

    # --- 噪声 ---
    noise_vals = _values(room, 'noise')
    noise_result = _result(room, 'noise')
    if noise_vals or noise_result:
        judge_range = _get_judge_range(export_payload, 'noise')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'噪声（dB(A)）{title_suffix}'
        row = _write_section_simple(ws, row, title, noise_vals or [noise_result], noise_result, _judge_text(room, 'noise'))

    # --- 照度（最低照度） ---
    illum_vals = _values(room, 'illumination', 'illumination_main_room', 'illumination_min')
    if illum_vals:
        judge_range = _get_judge_range(export_payload, 'illumination')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'照度（最低照度）（lx）{title_suffix}'
        row = _write_section_simple(ws, row, title, illum_vals, _result(room, 'illumination', 'illumination_main_room', 'illumination_min'), _judge_text(room, 'illumination', 'illumination_main_room', 'illumination_min'))

    # --- 照度均匀度 ---
    illum_uni_vals = _values(room, 'illumination_uniformity')
    if illum_uni_vals:
        judge_range = _get_judge_range(export_payload, 'illumination_uniformity')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'照度均匀度{title_suffix}'
        row = _write_section_simple(ws, row, title, illum_uni_vals, _result(room, 'illumination_uniformity'), _judge_for_key(_result(room, 'illumination_uniformity'), 'illumination_uniformity'))

    # --- 工作照度（动物房） ---
    work_illum_vals = _values(room, 'work_illumination')
    work_illum_result = _result(room, 'work_illumination')
    if work_illum_vals or work_illum_result:
        judge_range = _get_judge_range(export_payload, 'work_illumination')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'工作照度（lx）{title_suffix}'
        row = _write_section_simple(ws, row, title, work_illum_vals or [work_illum_result], work_illum_result, _judge_for_key(work_illum_result, 'work_illumination'))

    # --- 动物照度（动物房） ---
    animal_illum_vals = _values(room, 'animal_illumination')
    animal_illum_result = _result(room, 'animal_illumination')
    if animal_illum_vals or animal_illum_result:
        judge_range = _get_judge_range(export_payload, 'animal_illumination')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'动物照度（lx）{title_suffix}'
        row = _write_section_simple(ws, row, title, animal_illum_vals or [animal_illum_result], animal_illum_result, _judge_for_key(animal_illum_result, 'animal_illumination'))

    # --- 温差（动物房） ---
    temp_diff_vals = _values(room, 'temp_diff')
    temp_diff_result = _result(room, 'temp_diff')
    if temp_diff_vals or temp_diff_result:
        judge_range = _get_judge_range(export_payload, 'temp_diff')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'日温差（℃）{title_suffix}'
        row = _write_section_simple(ws, row, title, temp_diff_vals or [temp_diff_result], temp_diff_result, _judge_for_key(temp_diff_result, 'temp_diff'))

    # --- 气流方向（负压病房） ---
    airflow_dir_vals = _values(room, 'airflow_direction')
    airflow_dir_result = _result(room, 'airflow_direction')
    if airflow_dir_vals or airflow_dir_result:
        judge_range = _get_judge_range(export_payload, 'airflow_direction')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'气流方向{title_suffix}'
        row = _write_section_simple(ws, row, title, airflow_dir_vals or [airflow_dir_result], airflow_dir_result, _judge_for_key(airflow_dir_result, 'airflow_direction'))

    # --- 沉降菌 ---
    settling_vals = _values(room, 'settling', 'settle_bacteria')
    settling_data = _data(room, 'settling', 'settle_bacteria')
    if settling_vals or settling_data:
        judge_range = _get_judge_range(export_payload, 'settling')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'沉降菌（cfu/皿）{title_suffix}'
        raw = settling_vals or ([str(settling_data.get('total'))] if settling_data.get('total') not in (None, '') else [''])
        result = _result(room, 'settling', 'settle_bacteria')
        if not result and settling_data:
            total = settling_data.get('total', '')
            blank = settling_data.get('blank', '0')
            neg = settling_data.get('neg', '0')
            result = f'平均值:{total} | 空白对照:{blank} | 阴性对照:{neg}'
        row = _write_section_simple(ws, row, title, raw, result, _judge_for_key(_get_param(room, 'settling', 'settle_bacteria').get('result', '') or result, 'settling'))

    # --- 浮游菌 ---
    floating_vals = _values(room, 'floating', 'floating_bacteria')
    floating_data = _data(room, 'floating', 'floating_bacteria')
    if floating_vals or floating_data:
        judge_range = _get_judge_range(export_payload, 'floating')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'浮游菌（cfu/m³）{title_suffix}'
        raw = floating_vals or ([str(floating_data.get('total'))] if floating_data.get('total') not in (None, '') else [''])
        result = _result(room, 'floating', 'floating_bacteria')
        if not result and floating_data:
            total = floating_data.get('total', '')
            blank = floating_data.get('blank', '0')
            neg = floating_data.get('neg', '0')
            result = f'平均浓度:{total} | 空白对照:{blank} | 阴性对照:{neg}'
        row = _write_section_simple(ws, row, title, raw, result, _judge_for_key(_get_param(room, 'floating', 'floating_bacteria').get('result', '') or result, 'floating'))

    # --- 表面菌（负压病房） ---
    surface_vals = _values(room, 'surface_bacteria')
    surface_result = _result(room, 'surface_bacteria')
    if surface_vals or surface_result:
        judge_range = _get_judge_range(export_payload, 'surface_bacteria')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'表面菌（cfu/cm²）{title_suffix}'
        row = _write_section_simple(ws, row, title, surface_vals or [surface_result], surface_result, _judge_for_key(surface_result, 'surface_bacteria'))

    # 签名行
    row = _write_signature(ws, row + 1, project)

    wb.save(output)
    wb.close()
    return str(output)

# ============================================================
# 制药/食品洁净车间类检测项导出
# ============================================================
def _build_pharma_cleanroom_record(export_payload, output_path):
    """
    制药类：gmp_workshop, veterinary_gmp_workshop, food_workshop
    检测项：换气次数、静压差、高效过滤器检漏、洁净度（悬浮粒子）、温度、相对湿度、
           噪声、照度、自净时间、沉降菌、浮游菌
    """
    global _current_judgement_index
    _current_judgement_index = _build_judgement_index(export_payload)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    project = export_payload.get('project', {}) or {}
    room = export_payload.get('room', {}) or {}
    ws.title = str(room.get('room_name') or '原始记录')[:31]
    _set_base_layout(ws)

    row = _write_page_header(ws, project, room, export_payload)

    # --- 换气次数 ---
    vents = _vents(room, 'airchange')
    airchange_vals = _values(room, 'airchange')
    if vents or airchange_vals:
        judge_range = _get_judge_range(export_payload, 'airchange')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'换气次数（次/h）{title_suffix}'
        if vents:
            v = vents[0]
            area = float(v.get('area') or 0)
            speed = float(v.get('speed') or 0)
            volume = float(v.get('volume') or 0)
            length = float(room.get('length') or 0)
            width = float(room.get('width') or 0)
            height = float(room.get('height') or 0)
            room_volume = round(length * width * height, 1) if length and width and height else 0
            if not volume and area and speed:
                volume = round(area * speed * 3600, 1)
            headers = ['风口面积(m²)', '风速(m/s)', '风量(m³/h)', '房间体积(m³)', '换气次数', '', '', '']
            data = [str(area), str(speed), str(volume), str(room_volume), _result(room, 'airchange'), '', '', '']
            row = _write_section(ws, row, title, headers, [data], _result(room, 'airchange'), _judge_text(room, 'airchange'))
        else:
            row = _write_section_simple(ws, row, title, airchange_vals, _result(room, 'airchange'), _judge_text(room, 'airchange'))

    # --- 截面平均风速（兽药车间） ---
    wind_vals = _values(room, 'wind_speed', 'air_velocity')
    wind_result = _result(room, 'wind_speed', 'air_velocity')
    if wind_vals or wind_result:
        judge_range = _get_judge_range(export_payload, 'wind_speed')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'截面平均风速（m/s）{title_suffix}'
        row = _write_section_simple(ws, row, title, wind_vals or [wind_result], wind_result, _judge_for_key(wind_result, 'wind_speed'))

    # --- 风速不均匀度（兽药车间） ---
    uniformity_vals = _values(room, 'wind_speed_uniformity', 'wind_uniformity', 'velocity_uniformity')
    uniformity_result = _result(room, 'wind_speed_uniformity', 'wind_uniformity', 'velocity_uniformity')
    if uniformity_vals or uniformity_result:
        judge_range = _get_judge_range(export_payload, 'wind_speed_uniformity')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'风速不均匀度{title_suffix}'
        row = _write_section_simple(ws, row, title, uniformity_vals or [uniformity_result], uniformity_result, _judge_for_key(uniformity_result, 'wind_speed_uniformity'))

    # --- 静压差 ---
    prs = _pairs(room, 'pressure')
    pressure_vals = _values(room, 'pressure')
    if prs or pressure_vals:
        judge_range = _get_judge_range(export_payload, 'pressure')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'静压差（Pa）{title_suffix}'
        if prs:
            headers = ['相对区域', '测点1', '测点2', '测点3', '平均值', '', '', '']
            data_rows = []
            for pair in prs:
                ref = pair.get('ref', '') or pair.get('reference', '')
                vals = pair.get('values') or []
                avg = pair.get('avg', '') or pair.get('result', '')
                row_data = [ref] + [str(v) for v in vals[:3]] + [str(avg)] + [''] * 3
                data_rows.append(row_data[:8])
            row = _write_section(ws, row, title, headers, data_rows, _result(room, 'pressure'), _judge_for_key(_result(room, 'pressure'), 'pressure'))
        else:
            row = _write_section_simple(ws, row, title, pressure_vals, _result(room, 'pressure'), _judge_for_key(_result(room, 'pressure'), 'pressure'))

    # --- 高效过滤器检漏 ---
    hepa = _get_param(room, 'hepa_leak')
    if hepa:
        judge_range = _get_judge_range(export_payload, 'hepa_leak')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'高效过滤器检漏（%）{title_suffix}'
        hepa_vals = _values(room, 'hepa_leak')
        if not hepa_vals:
            total = _data(room, 'hepa_leak').get('total', '')
            hepa_vals = [str(total)] if total not in (None, '') else ['']
        row = _write_section_simple(ws, row, title, hepa_vals, _result(room, 'hepa_leak'), _judge_for_key(_result(room, 'hepa_leak'), 'hepa_leak'))

    # --- 洁净度（悬浮粒子） ---
    pdata = _data(room, 'particle')
    particle_vals = _values(room, 'particle')
    particle_param = _get_param(room, 'particle')
    if pdata or particle_vals or particle_param:
        judge_range = _get_judge_range(export_payload, 'particle')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'洁净度（悬浮粒子）（粒/m³）{title_suffix}'

        if pdata:
            zones_data = [{'zone': '检测区域', 'particles': []}]
            if pdata.get('p05_max') or pdata.get('p05_ucl'):
                zones_data[0]['particles'].append({'size': '≥0.5μm', 'max': pdata.get('p05_max', ''), 'ucl': pdata.get('p05_ucl', '')})
            if pdata.get('p5_max') or pdata.get('p5_ucl'):
                zones_data[0]['particles'].append({'size': '≥5μm', 'max': pdata.get('p5_max', ''), 'ucl': pdata.get('p5_ucl', '')})
            if zones_data[0]['particles']:
                row = _write_cleanness_section(ws, row, title, zones_data, _result(room, 'particle'), _judge_for_key(particle_param.get('result', ''), 'particle'))
            else:
                row = _write_section_simple(ws, row, title, particle_vals, _result(room, 'particle'), _judge_for_key(particle_param.get('result', ''), 'particle'))
        else:
            row = _write_section_simple(ws, row, title, particle_vals, _result(room, 'particle'), _judge_for_key(particle_param.get('result', ''), 'particle'))

    # --- 温度 ---
    temp_vals = _values(room, 'temperature')
    if temp_vals:
        judge_range = _get_judge_range(export_payload, 'temperature')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'温度（℃）{title_suffix}'
        row = _write_section_simple(ws, row, title, temp_vals, _result(room, 'temperature'), _judge_text(room, 'temperature'))

    # --- 相对湿度 ---
    humid_vals = _values(room, 'humidity')
    if humid_vals:
        judge_range = _get_judge_range(export_payload, 'humidity')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'相对湿度（%）{title_suffix}'
        row = _write_section_simple(ws, row, title, humid_vals, _result(room, 'humidity'), _judge_text(room, 'humidity'))

    # --- 噪声 ---
    noise_vals = _values(room, 'noise')
    noise_result = _result(room, 'noise')
    if noise_vals or noise_result:
        judge_range = _get_judge_range(export_payload, 'noise')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'噪声（dB(A)）{title_suffix}'
        row = _write_section_simple(ws, row, title, noise_vals or [noise_result], noise_result, _judge_text(room, 'noise'))

    # --- 照度 ---
    illum_vals = _values(room, 'illumination', 'illumination_main_room')
    if illum_vals:
        judge_range = _get_judge_range(export_payload, 'illumination')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'照度（lx）{title_suffix}'
        row = _write_section_simple(ws, row, title, illum_vals, _result(room, 'illumination', 'illumination_main_room'), _judge_text(room, 'illumination', 'illumination_main_room'))

    # --- 自净时间 ---
    self_purify_vals = _values(room, 'self_purification_time', 'self_purify_time')
    self_purify_result = _result(room, 'self_purification_time', 'self_purify_time')
    if self_purify_vals or self_purify_result:
        judge_range = _get_judge_range(export_payload, 'self_purification_time')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'自净时间（min）{title_suffix}'
        row = _write_section_simple(ws, row, title, self_purify_vals or [self_purify_result], self_purify_result, _judge_for_key(self_purify_result, 'self_purification_time'))

    # --- 气流流型 ---
    airflow_vals = _values(room, 'airflow_pattern')
    airflow_result = _result(room, 'airflow_pattern')
    if airflow_vals or airflow_result:
        judge_range = _get_judge_range(export_payload, 'airflow_pattern')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'气流流型{title_suffix}'
        row = _write_section_simple(ws, row, title, airflow_vals or [airflow_result], airflow_result, _judge_for_key(airflow_result, 'airflow_pattern'))

    # --- 沉降菌 ---
    settling_vals = _values(room, 'settling', 'settle_bacteria')
    settling_data = _data(room, 'settling', 'settle_bacteria')
    if settling_vals or settling_data:
        judge_range = _get_judge_range(export_payload, 'settling')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'沉降菌（cfu/皿）{title_suffix}'
        raw = settling_vals or ([str(settling_data.get('total'))] if settling_data.get('total') not in (None, '') else [''])
        result = _result(room, 'settling', 'settle_bacteria')
        if not result and settling_data:
            total = settling_data.get('total', '')
            blank = settling_data.get('blank', '0')
            neg = settling_data.get('neg', '0')
            result = f'平均值:{total} | 空白对照:{blank} | 阴性对照:{neg}'
        row = _write_section_simple(ws, row, title, raw, result, _judge_for_key(_get_param(room, 'settling', 'settle_bacteria').get('result', '') or result, 'settling'))

    # --- 浮游菌 ---
    floating_vals = _values(room, 'floating', 'floating_bacteria')
    floating_data = _data(room, 'floating', 'floating_bacteria')
    if floating_vals or floating_data:
        judge_range = _get_judge_range(export_payload, 'floating')
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'浮游菌（cfu/m³）{title_suffix}'
        raw = floating_vals or ([str(floating_data.get('total'))] if floating_data.get('total') not in (None, '') else [''])
        result = _result(room, 'floating', 'floating_bacteria')
        if not result and floating_data:
            total = floating_data.get('total', '')
            blank = floating_data.get('blank', '0')
            neg = floating_data.get('neg', '0')
            result = f'平均浓度:{total} | 空白对照:{blank} | 阴性对照:{neg}'
        row = _write_section_simple(ws, row, title, raw, result, _judge_for_key(_get_param(room, 'floating', 'floating_bacteria').get('result', '') or result, 'floating'))

    # 签名行
    row = _write_signature(ws, row + 1, project)

    wb.save(output)
    wb.close()
    return str(output)

# ============================================================
# 电子车间类检测项导出
# ============================================================
def _build_electronics_record(export_payload, output_path):
    """
    电子车间：electronics_workshop
    与制药类类似，但可能有不同的检测项组合
    """
    # 电子车间使用与制药类相同的结构
    return _build_pharma_cleanroom_record(export_payload, output_path)


# ============================================================
# Fallback 通用导出
# ============================================================
def _build_fallback_record(export_payload, output_path):
    """通用 fallback：遍历所有 params 导出"""
    global _current_judgement_index
    _current_judgement_index = _build_judgement_index(export_payload)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    project = export_payload.get('project', {}) or {}
    room = export_payload.get('room', {}) or {}
    ws.title = str(room.get('room_name') or room.get('type_name') or '原始记录')[:31]
    _set_base_layout(ws)

    row = _write_page_header(ws, project, room, export_payload)

    # 遍历所有参数
    params = room.get('params', {}) or {}
    if isinstance(params, dict):
        iterable = list(params.items())
    elif isinstance(params, list):
        iterable = []
        for idx, item in enumerate(params, start=1):
            if isinstance(item, dict):
                key = item.get('key') or item.get('label') or item.get('name') or f'param_{idx}'
            else:
                key = f'param_{idx}'
            iterable.append((key, item))
    else:
        iterable = []

    for key, item in iterable:
        if not isinstance(item, dict):
            item = {'result': item}
        result = _clean_text(item.get('result', ''))
        raw_values = item.get('values')
        if isinstance(raw_values, list):
            vals = [str(v) for v in raw_values if str(v).strip()]
        elif raw_values is not None:
            vals = [str(raw_values)]
        elif item.get('value') is not None:
            vals = [str(item.get('value'))]
        else:
            vals = []

        judge_range = _get_judge_range(export_payload, key)
        title_suffix = f' 判定范围：{judge_range}' if judge_range else ''
        title = f'{key}{title_suffix}'
        row = _write_section_simple(ws, row, title, vals or [''], result, _judge_for_key(item.get('result', ''), key))

    # 签名行
    row = _write_signature(ws, row + 1, project)

    wb.save(output)
    wb.close()
    return str(output)


# ============================================================
# 入口函数（签名不变）
# ============================================================
def build_canonical_excel_report(export_payload: dict, output_path: str):
    """
    主入口函数。
    多房间时每个房间一个Sheet，单房间直接导出。
    """
    rooms_export = export_payload.get('rooms_export') or []
    if len(rooms_export) <= 1:
        # 单房间：如果顶层没有 room 但 rooms_export 有一个，提升到顶层
        if rooms_export and not export_payload.get('room'):
            virtual = dict(export_payload)
            virtual['room'] = rooms_export[0].get('room', {})
            virtual['template_rule'] = rooms_export[0].get('template_rule', {})
            virtual['template_resource'] = rooms_export[0].get('template_resource', {})
            virtual['report_context'] = rooms_export[0].get('report_context', {})
            virtual['clean_class_semantics'] = rooms_export[0].get('clean_class_semantics', {})
            virtual['judgement_result'] = rooms_export[0].get('judgement_result', {})
            return _build_single_room_excel(virtual, output_path)
        return _build_single_room_excel(export_payload, output_path)

    # 多房间：每个房间一个 Sheet
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    default_ws = wb.active

    for i, room_export in enumerate(rooms_export):
        virtual_payload = dict(export_payload)
        virtual_payload['room'] = room_export.get('room', {})
        virtual_payload['template_rule'] = room_export.get('template_rule', {})
        virtual_payload['template_resource'] = room_export.get('template_resource', {})
        virtual_payload['report_context'] = room_export.get('report_context', {})
        virtual_payload['clean_class_semantics'] = room_export.get('clean_class_semantics', {})
        virtual_payload['judgement_result'] = room_export.get('judgement_result', {})

        room_name = room_export.get('room', {}).get('room_name', '') or f'房间{i+1}'
        sheet_name = str(room_name)[:31]

        # 导出到临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        _build_single_room_excel(virtual_payload, tmp_path)

        # 读取临时文件并复制到主 workbook
        src_wb = openpyxl.load_workbook(tmp_path)
        src_ws = src_wb.active

        if i == 0:
            ws = default_ws
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # 复制单元格内容和格式
        for src_row in src_ws.iter_rows():
            for cell in src_row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.number_format = cell.number_format

        # 复制合并单元格
        for merged_range in src_ws.merged_cells.ranges:
            ws.merge_cells(str(merged_range))

        # 复制列宽
        for col_letter, col_dim in src_ws.column_dimensions.items():
            ws.column_dimensions[col_letter].width = col_dim.width

        # 复制行高
        for row_idx, row_dim in src_ws.row_dimensions.items():
            if row_dim.height:
                ws.row_dimensions[row_idx].height = row_dim.height

        src_wb.close()
        os.unlink(tmp_path)

    wb.save(output)
    wb.close()
    return str(output)


def _build_single_room_excel(export_payload: dict, output_path: str):
    """单房间导出，按 type_id 分发到对应的构建函数"""
    room = export_payload.get('room', {}) or {}
    type_id = str(room.get('type_id', '') or '')

    # 制药/食品洁净车间
    if type_id in {'gmp_workshop', 'veterinary_gmp_workshop', 'food_workshop'}:
        return _build_pharma_cleanroom_record(export_payload, output_path)

    # 医院类
    if type_id in {'negative_pressure', 'operating_room', 'clean_function_room', 'bsl', 'animal_room'}:
        return _build_hospital_record(export_payload, output_path)

    # 电子车间
    if type_id in {'electronics_workshop'}:
        return _build_electronics_record(export_payload, output_path)

    # 其他 fallback
    return _build_fallback_record(export_payload, output_path)
