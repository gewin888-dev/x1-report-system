#!/usr/bin/env python3
"""
X1 检测记录系统 - 独立运行骨架

设计目标：
1. 与 V/T 完全隔离
2. 独立目录、独立记录、独立报告、独立日志、独立缓存
3. 先提供 X1 自己的首页与 API 骨架，不复用 T/V 的运行态路径
"""

from pathlib import Path
import json
import os
import re
import sqlite3
import secrets
import psutil
import subprocess
import shutil
import tarfile
import tempfile
import time
from datetime import datetime
from collections import defaultdict
from flask import Flask, render_template, jsonify, request, redirect, url_for, flash, session, send_file
from flask_compress import Compress
from flask_login import login_user, logout_user, login_required, current_user
from adapters.export_docx import build_canonical_object_report
from adapters.export_excel import build_canonical_excel_report
from adapters.template_fill import build_template_bound_docx, build_mixed_report_docx
from template_rules import resolve_template_rule
from report_context_builder import build_report_context
from clean_class_semantics import build_clean_class_semantics, _normalize_operating_room_context
from judgement_engine import judge_room
from template_resources import resolve_template_resource, apply_type_default_template, apply_semantic_default_template
from config_loader import load_x1_config
from feishu_utils import resolve_feishu_upload_folder, get_feishu_folder_meta, upload_file_to_feishu, get_feishu_config, get_feishu_token, download_file_from_feishu, download_file_content_from_feishu
from payload_normalizer import normalize_project_payload, validate_normalized_project
from database import init_database, get_db
from auth import init_login_manager, get_user, verify_password, require_role, require_permission, DEFAULT_ROLE_PERMISSIONS, can_view_record
from monitor import log_action, log_error, monitor_performance, get_system_health
from notifications import (create_notification, get_notifications, get_unread_count,
    mark_read, mark_all_read, ensure_notifications_table,
    notify_new_registration, notify_customer_urge, notify_report_feedback,
    notify_report_ready, notify_registration_approved, notify_registration_rejected)
from helpers.record_utils import _compute_record_asset_state, _soft_delete_record, _can_access_file_by_name, cleanup_trash, _x_now, _x_draft_path
from helpers.settings_utils import _setting_enabled, _load_system_settings
BASE_DIR = Path(__file__).parent
CFG = load_x1_config(BASE_DIR)
APP_VERSION = CFG.get('version', 'UNKNOWN_VERSION')
APP_PORT = int(CFG.get('port', 8082))
APP_HOST = CFG.get('host', '127.0.0.1')
HOST_MODE = str(CFG.get('host_mode', 'desktop') or 'desktop').strip().lower()
ALLOWED_SETTINGS_BROWSE_ROOTS = [Path(BASE_DIR), Path.home()]

PATHS = CFG.get('paths', {})
FORMAL_RECORDS_BASE = Path(os.path.expanduser(str((CFG.get('archive') or {}).get('formal_raw_archive') or PATHS.get('formal_raw_archive') or '~/公司资料/检测部/原始记录'))).resolve()
FORMAL_REPORTS_BASE = Path(os.path.expanduser(str((CFG.get('archive') or {}).get('formal_report_archive') or PATHS.get('formal_report_archive') or '~/公司资料/检测部/检测报告'))).resolve()

LOG_ACTION_CATEGORY_MAP = {
    'login': '认证与会话',
    'login_failed': '认证与会话',
    'logout': '认证与会话',
    '导出报告': '报告记录治理',
    '删除记录': '报告记录治理',
    '批量删除记录': '报告记录治理',
    '作废记录': '报告记录治理',
    '重试飞书上传': '报告记录治理',
    '批量删除操作日志': '操作日志治理',
    '更新系统设置': '系统设置与运维',
    '执行路径巡检': '系统设置与运维',
    '创建系统设置目录': '系统设置与运维',
    '创建子目录': '系统设置与运维',
    '测试飞书配置': '系统设置与运维',
    '执行立即备份': '系统设置与运维',
    '更新角色权限': '权限与用户管理',
    '新增用户': '权限与用户管理',
    '编辑用户': '权限与用户管理',
    '删除用户': '权限与用户管理',
    '重置用户密码': '权限与用户管理',
    '注册模板': '模板注册与模板治理',
    '上传并注册模板': '模板注册与模板治理',
    '模板导出验证': '模板注册与模板治理',
    '删除模板注册': '模板注册与模板治理',
    '切换模板启停': '模板注册与模板治理',
    '验证模板': '模板注册与模板治理',
    '设置默认模板': '模板注册与模板治理',
    '加入模板候选': '模板注册与模板治理',
    '设置语义默认模板': '模板注册与模板治理',
    '加入语义模板候选': '模板注册与模板治理',
    '查看模板': '标准库/文档/查看类',
}

LOG_ACTION_CATEGORIES = [
    '认证与会话', '报告记录治理', '操作日志治理', '系统设置与运维',
    '权限与用户管理', '模板注册与模板治理', '标准库/文档/查看类', '其他/未分类'
]

TASK_TYPE_OPTIONS = [
    'inspection',
    'reinspection',
    'supplementary',
    'rectification',
]

TASK_STATUS_OPTIONS = [
    'pending_assign',
    'assigned',
    'accepted',
    'in_progress',
    'completed',
    'cancelled',
]

TASK_TYPE_LABELS = {
    'inspection': '常规检测',
    'reinspection': '复检任务',
    'supplementary': '补测任务',
    'rectification': '整改任务',
}

TASK_STATUS_LABELS = {
    'pending_assign': '待指派',
    'assigned': '已派单',
    'accepted': '已接单',
    'in_progress': '执行中',
    'completed': '已完成',
    'cancelled': '已取消',
}


def get_log_action_category(action: str) -> str:
    return LOG_ACTION_CATEGORY_MAP.get((action or '').strip(), '其他/未分类')


def _resolve_browse_path(path_str: str) -> Path:
    raw = (path_str or '').strip()
    candidate = Path(raw).expanduser().resolve() if raw else BASE_DIR.resolve()
    for root in ALLOWED_SETTINGS_BROWSE_ROOTS:
        try:
            candidate.relative_to(root.resolve())
            return candidate
        except Exception:
            continue
    return ALLOWED_SETTINGS_BROWSE_ROOTS[0].resolve()


def _is_desktop_mode() -> bool:
    return HOST_MODE == 'desktop'


# 模板基础路径：从配置文件读取，支持 ~ 展开
template_base_config = CFG.get('template_base', '~/公司资料/检测部/检测报告模板')
if template_base_config.startswith('~'):
    TEMPLATE_BASE = Path.home() / template_base_config[2:]  # 去掉 ~/
else:
    TEMPLATE_BASE = Path(template_base_config)
TEMPLATE_BASE = TEMPLATE_BASE.expanduser().resolve()

TEMPLATE_MAP_X1 = [
    ('hospital', 'operating_room', 'Ⅰ级', '医院洁净部/医院洁净部洁净手术部手术室百级检测报告模板.docx'),
    ('hospital', 'operating_room', 'Ⅱ级', '医院洁净部/医院洁净部洁净手术部手术室千级检测报告模板.docx'),
    ('hospital', 'operating_room', 'Ⅲ级', '医院洁净部/医院洁净部洁净手术部手术室万级检测报告模板.docx'),
    ('hospital', 'operating_room', 'Ⅳ级', '医院洁净部/医院洁净部洁净手术部手术室十万级检测报告模板.docx'),
]

PATHS = CFG.get('paths', {})
STATIC_DIR = BASE_DIR / PATHS.get('static', 'static')
TEMPLATES_DIR = BASE_DIR / PATHS.get('templates', 'templates')
RECORDS_DIR = BASE_DIR / PATHS.get('records', 'records_x1')
REPORTS_DIR = BASE_DIR / PATHS.get('reports', 'reports_x1')
LOGS_DIR = BASE_DIR / PATHS.get('logs', 'logs_x1')
CACHE_DIR = BASE_DIR / PATHS.get('cache', 'cache_x1')
TEMP_DIR = BASE_DIR / PATHS.get('temp', 'temp_x1')
UPLOADS_DIR = BASE_DIR / PATHS.get('uploads', 'uploads_x1')
ADAPTERS_DIR = BASE_DIR / PATHS.get('adapters', 'adapters')
DOCS_DIR = BASE_DIR / PATHS.get('docs', 'docs')
TEMPLATE_CONFIG_FILE = BASE_DIR / 'template_config.json'

FORMAL_RECORDS_BASE = Path(os.path.expanduser(str((CFG.get('archive') or {}).get('formal_raw_archive') or PATHS.get('formal_raw_archive') or '~/公司资料/检测部/原始记录'))).resolve()
FORMAL_REPORTS_BASE = Path(os.path.expanduser(str((CFG.get('archive') or {}).get('formal_report_archive') or PATHS.get('formal_report_archive') or '~/公司资料/检测部/检测报告'))).resolve()

def _formal_year_dir(base: Path, year: int) -> Path:
    target = base / str(year)
    target.mkdir(parents=True, exist_ok=True)
    return target

def _safe_filename_part(value: str, fallback: str = '未命名') -> str:
    text = str(value or '').strip()
    if not text:
        text = fallback
    for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        text = text.replace(ch, '_')
    text = text.replace('\n', '_').replace('\r', '_').replace('\t', '_')
    return text[:120].strip() or fallback


def _copy_to_formal_dir(src: Path, base: Path, year: int, target_name=None):
    if not src or not src.exists():
        return {'success': False, 'error': '源文件不存在'}
    try:
        import shutil
        target_dir = _formal_year_dir(base, year)
        target = target_dir / (target_name or src.name)
        shutil.copy2(src, target)
        return {'success': True, 'path': str(target), 'filename': target.name}
    except Exception as e:
        return {'success': False, 'error': str(e)}

for p in [RECORDS_DIR, REPORTS_DIR, LOGS_DIR, CACHE_DIR, TEMP_DIR, UPLOADS_DIR, ADAPTERS_DIR, DOCS_DIR]:
    p.mkdir(parents=True, exist_ok=True)

SECRET_FILE = BASE_DIR / 'data' / 'flask_secret_key.txt'

def _load_or_create_secret_key():
    env_secret = os.getenv('X1_SECRET_KEY', '').strip()
    if env_secret:
        return env_secret
    if SECRET_FILE.exists():
        try:
            existing = SECRET_FILE.read_text(encoding='utf-8').strip()
            if existing:
                return existing
        except Exception:
            pass
    secret = secrets.token_urlsafe(48)
    SECRET_FILE.parent.mkdir(parents=True, exist_ok=True)
    SECRET_FILE.write_text(secret, encoding='utf-8')
    return secret


def get_x1_data_conn():
    db_path = BASE_DIR / 'data' / 'x1_data.db'
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn


def init_business_projects_table():
    conn = get_x1_data_conn()
    try:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS business_projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT NOT NULL,
            client_name TEXT DEFAULT '',
            project_address TEXT DEFAULT '',
            contact_name TEXT DEFAULT '',
            contact_phone TEXT DEFAULT '',
            detection_domain TEXT DEFAULT '',
            detection_type TEXT DEFAULT '',
            expected_detection_date TEXT DEFAULT '',
            project_desc TEXT DEFAULT '',
            business_stage TEXT DEFAULT '',
            contract_status TEXT DEFAULT '',
            contract_amount REAL DEFAULT 0,
            inspection_stage TEXT DEFAULT '',
            report_status TEXT DEFAULT '',
            invoice_status TEXT DEFAULT '',
            payment_status TEXT DEFAULT '',
            owner TEXT DEFAULT '',
            remarks TEXT DEFAULT '',
            assigned_to TEXT DEFAULT '',
            assigned_at TEXT DEFAULT '',
            task_status TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
        """)
        # 增量迁移：项目编号字段
        try:
            conn.execute("ALTER TABLE business_projects ADD COLUMN project_no TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass  # 已存在
        # 增量迁移：已收款金额
        try:
            conn.execute("ALTER TABLE business_projects ADD COLUMN paid_amount REAL DEFAULT 0")
        except sqlite3.OperationalError:
            pass
        # 增量迁移：报告文件路径（补录场景）
        try:
            conn.execute("ALTER TABLE business_projects ADD COLUMN report_file_path TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        conn.commit()
    finally:
        conn.close()


def _generate_project_no():
    """生成项目编号，格式：PJ-YYYY-NNNN"""
    year = datetime.now().strftime('%Y')
    prefix = f'PJ-{year}-'
    conn = get_x1_data_conn()
    try:
        row = conn.execute(
            "SELECT project_no FROM business_projects "
            "WHERE project_no LIKE ? ORDER BY project_no DESC LIMIT 1",
            (f'{prefix}%',)
        ).fetchone()
        if row and row['project_no']:
            try:
                last_seq = int(row['project_no'].replace(prefix, ''))
            except ValueError:
                last_seq = 0
        else:
            last_seq = 0
        return f'{prefix}{last_seq + 1:04d}'
    finally:
        conn.close()


def serialize_business_project(row):
    if not row:
        return None
    return {
        'id': row['id'],
        'project_no': row['project_no'] or '',
        'project_name': row['project_name'] or '',
        'client_name': row['client_name'] or '',
        'project_address': row['project_address'] or '',
        'contact_name': row['contact_name'] or '',
        'contact_phone': row['contact_phone'] or '',
        'detection_domain': row['detection_domain'] or '',
        'detection_type': row['detection_type'] or '',
        'expected_detection_date': row['expected_detection_date'] or '',
        'project_desc': row['project_desc'] or '',
        'business_stage': row['business_stage'] or '',
        'contract_status': row['contract_status'] or '',
        'contract_amount': row['contract_amount'] or 0,
        'paid_amount': row['paid_amount'] if 'paid_amount' in row.keys() else 0,
        'receivable_amount': round((row['contract_amount'] or 0) - (row['paid_amount'] if 'paid_amount' in row.keys() else 0), 2),
        'inspection_stage': row['inspection_stage'] or '',
        'report_status': row['report_status'] or '',
        'invoice_status': row['invoice_status'] or '',
        'payment_status': row['payment_status'] or '',
        'owner': row['owner'] or '',
        'remarks': row['remarks'] or '',
        'assigned_to': row['assigned_to'] or '',
        'assigned_at': row['assigned_at'] or '',
        'task_status': row['task_status'] or '',
        'created_at': row['created_at'] or '',
        'updated_at': row['updated_at'] or '',
        'source': row['source'] if 'source' in row.keys() else '',
        'has_urge': row['has_urge'] if 'has_urge' in row.keys() else '',
        'report_file_path': row['report_file_path'] if 'report_file_path' in row.keys() else '',
    }


def init_project_tasks_table():
    conn = get_x1_data_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                task_name TEXT NOT NULL,
                task_type TEXT NOT NULL DEFAULT 'inspection',
                assigned_to INTEGER,
                assigned_at TEXT,
                task_status TEXT NOT NULL DEFAULT 'pending_assign',
                expected_execute_date TEXT,
                started_at TEXT,
                completed_at TEXT,
                remarks TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        cur.execute("CREATE INDEX IF NOT EXISTS idx_project_tasks_project_id ON project_tasks(project_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_project_tasks_assigned_to ON project_tasks(assigned_to)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_project_tasks_status ON project_tasks(task_status)")
        conn.commit()
    finally:
        conn.close()


def _get_task_status_label(status):
    return TASK_STATUS_LABELS.get((status or '').strip(), (status or '').strip())


def _get_task_type_label(task_type):
    return TASK_TYPE_LABELS.get((task_type or '').strip(), (task_type or '').strip())


def _get_business_project_by_id(project_id):
    if not project_id:
        return None
    conn = get_x1_data_conn()
    try:
        return conn.execute("SELECT * FROM business_projects WHERE id=?", (project_id,)).fetchone()
    finally:
        conn.close()


def _get_user_display_name(user_id):
    if not user_id:
        return ''
    try:
        with get_db() as conn:
            row = conn.execute(
                "SELECT user_id, display_name FROM users WHERE user_id=?",
                (str(user_id),)
            ).fetchone()
            if not row:
                return ''
            return row['display_name'] or row['user_id'] or ''
    except Exception:
        return ''


def serialize_project_task(row, project_row=None):
    if not row:
        return None
    project_name = ''
    client_name = ''
    if project_row:
        project_name = project_row['project_name'] or ''
        client_name = project_row['client_name'] or ''
    assigned_to = row['assigned_to']
    created_by = row['created_by']
    return {
        'id': row['id'],
        'project_id': row['project_id'],
        'project_name': project_name,
        'client_name': client_name,
        'task_name': row['task_name'] or '',
        'task_type': row['task_type'] or '',
        'task_type_label': _get_task_type_label(row['task_type']),
        'assigned_to': assigned_to,
        'assigned_to_name': _get_user_display_name(assigned_to),
        'assigned_at': row['assigned_at'] or '',
        'task_status': row['task_status'] or '',
        'task_status_label': _get_task_status_label(row['task_status']),
        'expected_execute_date': row['expected_execute_date'] or '',
        'started_at': row['started_at'] or '',
        'completed_at': row['completed_at'] or '',
        'remarks': row['remarks'] or '',
        'created_by': created_by,
        'created_by_name': _get_user_display_name(created_by),
        'created_at': row['created_at'] or '',
        'updated_at': row['updated_at'] or '',
    }


def _clean_project_task_payload(data):
    data = data or {}

    def s(key):
        value = data.get(key)
        if value is None:
            return None
        value = str(value).strip()
        return value or None

    def i(key):
        value = data.get(key)
        if value in (None, ''):
            return None
        try:
            return int(value)
        except Exception:
            return None

    payload = {
        'project_id': i('project_id'),
        'task_name': s('task_name'),
        'task_type': s('task_type') or 'inspection',
        'assigned_to': s('assigned_to'),
        'expected_execute_date': s('expected_execute_date'),
        'task_status': s('task_status'),
        'remarks': s('remarks'),
    }

    if payload['task_type'] not in TASK_TYPE_OPTIONS:
        payload['task_type'] = 'inspection'
    if payload['task_status'] and payload['task_status'] not in TASK_STATUS_OPTIONS:
        payload['task_status'] = None
    return payload


def refresh_project_task_summary(project_id):
    if not project_id:
        return
    conn = get_x1_data_conn()
    try:
        rows = conn.execute(
            """
            SELECT * FROM project_tasks
            WHERE project_id=?
            ORDER BY
                CASE task_status
                    WHEN 'in_progress' THEN 1
                    WHEN 'accepted' THEN 2
                    WHEN 'assigned' THEN 3
                    WHEN 'pending_assign' THEN 4
                    WHEN 'completed' THEN 5
                    WHEN 'cancelled' THEN 6
                    ELSE 99
                END,
                updated_at DESC,
                id DESC
            """,
            (project_id,)
        ).fetchall()
        chosen = next((row for row in rows if row['task_status'] != 'cancelled'), None)
        if chosen:
            conn.execute(
                """
                UPDATE business_projects
                SET assigned_to=?, assigned_at=?, task_status=?, updated_at=?
                WHERE id=?
                """,
                (
                    str(chosen['assigned_to'] or ''),
                    chosen['assigned_at'] or '',
                    chosen['task_status'] or '',
                    datetime.now().isoformat(timespec='seconds'),
                    project_id,
                )
            )
        else:
            conn.execute(
                """
                UPDATE business_projects
                SET assigned_to='', assigned_at='', task_status='', updated_at=?
                WHERE id=?
                """,
                (datetime.now().isoformat(timespec='seconds'), project_id)
            )
        conn.commit()
    finally:
        conn.close()


# ============================================================
# 项目状态自动流转
# ============================================================

# 状态流转优先级（数值越大 = 越靠后）
_INSPECTION_STAGE_ORDER = {
    '未安排': 0, '': 0,
    '已排期': 1,
    '检测中': 2,
    '检测完成': 3,
    '检测异常': -1,  # 异常不自动流转
}
_REPORT_STATUS_ORDER = {
    '未开始': 0, '': 0,
    '编制中': 1, '审核中': 2, '待修改': 2, '待出具': 3,
    '已出具': 4, '待客户确认': 5, '客户已确认': 6, '已发送客户': 7,
}


def _auto_advance_project_stage(project_id, target_inspection=None, target_report=None):
    """自动推进项目状态，只前进不后退。
    
    规则：
    - 只在目标状态比当前状态“更靠后”时才更新
    - 异常状态不会被自动覆盖
    - 写操作日志
    """
    if not project_id:
        return
    conn = get_x1_data_conn()
    try:
        row = conn.execute("SELECT * FROM business_projects WHERE id=?", (project_id,)).fetchone()
        if not row:
            return
        
        updates = []
        params = []
        current_ins = (row['inspection_stage'] or '').strip()
        current_rpt = (row['report_status'] or '').strip()
        
        if target_inspection:
            cur_order = _INSPECTION_STAGE_ORDER.get(current_ins, -99)
            tgt_order = _INSPECTION_STAGE_ORDER.get(target_inspection, -99)
            # 只前进，且不覆盖异常状态
            if cur_order >= 0 and tgt_order > cur_order:
                updates.append('inspection_stage=?')
                params.append(target_inspection)
        
        if target_report:
            cur_order = _REPORT_STATUS_ORDER.get(current_rpt, -99)
            tgt_order = _REPORT_STATUS_ORDER.get(target_report, -99)
            if cur_order >= 0 and tgt_order > cur_order:
                updates.append('report_status=?')
                params.append(target_report)
        
        if updates:
            updates.append('updated_at=?')
            params.append(datetime.now().isoformat(timespec='seconds'))
            params.append(project_id)
            conn.execute(
                f"UPDATE business_projects SET {', '.join(updates)} WHERE id=?",
                params
            )
            conn.commit()
            log_action(
                getattr(current_user, 'id', 'system') if hasattr(current_user, 'id') else 'system',
                '项目状态自动流转',
                f'project_id={project_id}',
                f'inspection_stage: {current_ins}→{target_inspection or "-"}, report_status: {current_rpt}→{target_report or "-"}'
            )
    except Exception:
        pass
    finally:
        conn.close()


def _clean_project_payload(data):
    def s(key):
        return str(data.get(key, '') or '').strip()

    raw_amount = data.get('contract_amount', 0)
    try:
        contract_amount = float(raw_amount or 0)
    except Exception:
        contract_amount = 0.0

    raw_paid = data.get('paid_amount', 0)
    try:
        paid_amount = float(raw_paid or 0)
    except Exception:
        paid_amount = 0.0

    return {
        'project_name': s('project_name'),
        'client_name': s('client_name'),
        'project_address': s('project_address'),
        'contact_name': s('contact_name'),
        'contact_phone': s('contact_phone'),
        'detection_domain': s('detection_domain'),
        'detection_type': s('detection_type'),
        'expected_detection_date': s('expected_detection_date'),
        'project_desc': s('project_desc'),
        'business_stage': s('business_stage'),
        'contract_status': s('contract_status') or '未签',
        'contract_amount': contract_amount,
        'paid_amount': paid_amount,
        'inspection_stage': s('inspection_stage') or '未安排',
        'report_status': s('report_status') or '未开始',
        'invoice_status': s('invoice_status') or '未开票',
        'payment_status': s('payment_status') or '未回款',
        'owner': s('owner'),
        'remarks': s('remarks'),
        'assigned_to': s('assigned_to'),
        'assigned_at': s('assigned_at'),
        'task_status': s('task_status'),
    }

app = Flask(__name__, template_folder=str(TEMPLATES_DIR), static_folder=str(STATIC_DIR))
app.secret_key = CFG.get('secret_key') or _load_or_create_secret_key()
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['COMPRESS_MIMETYPES'] = ['text/html', 'text/css', 'application/json', 'application/javascript', 'text/javascript']
app.config['COMPRESS_MIN_SIZE'] = 500
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = bool(CFG.get('session_cookie_secure', False))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
Compress(app)

# 初始化认证系统
login_manager = init_login_manager(app)
init_business_projects_table()
init_project_tasks_table()

# 客户界面路由注册
from customer_routes import customer_bp, init_customer_tables
from customer_admin_routes import customer_admin_bp
init_customer_tables()
app.register_blueprint(customer_bp)
app.register_blueprint(customer_admin_bp)


@app.after_request
def add_no_cache_headers(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


def _same_origin_request() -> bool:
    origin = request.headers.get('Origin', '').strip()
    referer = request.headers.get('Referer', '').strip()
    host_url = request.host_url.rstrip('/')
    if origin:
        return origin.rstrip('/') == host_url
    if referer:
        return referer.startswith(host_url + '/') or referer == host_url
    return bool(request.is_json)


@app.before_request
def enforce_csrf_for_authenticated_writes():
    if request.method not in ('POST', 'PUT', 'PATCH', 'DELETE'):
        return None
    if not current_user.is_authenticated:
        return None
    if request.endpoint in {'login_page', 'admin_api_open_file', 'admin_api_open_feishu_file'}:
        return None
    if not _same_origin_request():
        return jsonify({'success': False, 'error': 'CSRF 校验失败'}), 403

@app.route('/static/manifest.json')
def x_manifest():
    return jsonify({
        'name': CFG.get('app_name', 'X1 检测记录系统'),
        'short_name': 'X1',
        'display': 'standalone',
        'start_url': '/',
        'background_color': '#ffffff',
        'theme_color': '#667eea',
        'icons': []
    })


def _compute_record_asset_state(record: dict) -> dict:
    files = record.get('files') or []
    report_info = record.get('report_info') or {}
    export_info = record.get('export_info') or {}
    report_file = next((f for f in files if '.filled.' in f.get('name', '')), None) or next((f for f in files if '.bound.' in f.get('name', '')), None)
    raw_excel = next((f for f in files if f.get('name', '').lower().endswith('.xlsx')), None)
    feishu_report_url = record.get('feishu_report_url') or report_info.get('feishu_url') or record.get('feishu_report_open_url') or report_info.get('feishu_open_url') or ''
    feishu_export_url = record.get('feishu_export_url') or export_info.get('feishu_url') or record.get('feishu_export_open_url') or export_info.get('feishu_open_url') or ''
    local_report_ok = bool(report_file)
    local_record_ok = bool(raw_excel)
    feishu_report_ok = bool(feishu_report_url) and record.get('feishu_report_status') != 'failed'
    feishu_record_ok = bool(feishu_export_url) and record.get('feishu_export_status') != 'failed'
    report_ready = bool(record.get('report_success') or record.get('has_report') or report_info.get('filename') or feishu_report_url or local_report_ok)
    raw_ready = bool(record.get('raw_record_success') or record.get('has_export') or export_info.get('filename') or feishu_export_url or local_record_ok)
    issues = []
    if record.get('type') == 'export' and not local_report_ok and not feishu_report_ok:
        issues.append('report_missing')
    if record.get('type') == 'export' and not local_record_ok and not feishu_record_ok:
        issues.append('raw_record_missing')
    if record.get('feishu_report_status') == 'failed':
        issues.append('feishu_report_failed')
    if record.get('feishu_export_status') == 'failed':
        issues.append('feishu_record_failed')
    return {
        'report_file': report_file,
        'raw_excel': raw_excel,
        'local_report_ok': local_report_ok,
        'local_record_ok': local_record_ok,
        'feishu_report_ok': feishu_report_ok,
        'feishu_record_ok': feishu_record_ok,
        'report_ready': report_ready,
        'raw_ready': raw_ready,
        'issues': issues,
        'healthy': len(issues) == 0
    }




@app.route('/api/user')
def api_user_compat():
    if current_user.is_authenticated:
        return jsonify({
            'username': current_user.id,
            'display_name': current_user.display_name,
            'role': current_user.role,
            'department': current_user.department,
            'permissions': sorted(list(current_user.permissions or []))
        })
    return jsonify({
        'username': 'guest',
        'display_name': '访客'
    })


# ==================== 登录限流防护 ====================
_login_attempts = defaultdict(list)  # {ip: [timestamp, ...]}
_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_WINDOW_SECONDS = 300  # 5分钟窗口
_LOGIN_LOCKOUT_SECONDS = 900  # 锁定15分钟

def _is_login_rate_limited(ip):
    """检查IP是否被限流"""
    now = time.time()
    attempts = _login_attempts[ip]
    # 清理过期记录
    _login_attempts[ip] = [t for t in attempts if now - t < _LOGIN_LOCKOUT_SECONDS]
    attempts = _login_attempts[ip]
    if len(attempts) >= _LOGIN_MAX_ATTEMPTS:
        oldest_in_window = [t for t in attempts if now - t < _LOGIN_WINDOW_SECONDS]
        if len(oldest_in_window) >= _LOGIN_MAX_ATTEMPTS:
            return True
    return False

def _record_login_failure(ip):
    """记录登录失败"""
    _login_attempts[ip].append(time.time())

def _clear_login_attempts(ip):
    """登录成功后清除记录"""
    _login_attempts.pop(ip, None)


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        client_ip = request.remote_addr or '0.0.0.0'
        
        # 暴力破解防护
        if _is_login_rate_limited(client_ip):
            msg = f'登录尝试过于频繁，请 {_LOGIN_LOCKOUT_SECONDS // 60} 分钟后再试'
            log_action('system', 'login_blocked', client_ip, f'IP限流触发: {client_ip}')
            if request.is_json:
                return jsonify({'success': False, 'error': msg}), 429
            flash(msg)
            return redirect(url_for('login_page'))
        
        # 支持JSON和表单两种方式
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form
        
        username = data.get('username')
        password = data.get('password')
        remember = data.get('remember', False)
        
        if verify_password(username, password):
            user = get_user(username)
            login_user(user, remember=remember)
            log_action(username, 'login', '', '登录成功')
            _clear_login_attempts(client_ip)
            
            if request.is_json:
                return jsonify({'success': True, 'redirect': '/home'})
            return redirect('/home')
        
        _record_login_failure(client_ip)
        log_action(username or 'unknown', 'login_failed', client_ip, f'密码错误 (IP: {client_ip})')
        if request.is_json:
            return jsonify({'success': False, 'error': '用户名或密码错误'}), 401
        flash('用户名或密码错误')
        return redirect(url_for('login_page'))
    
    return render_template('login.html', version=APP_VERSION)


@app.route('/customer/login')
def customer_login_page():
    """客户专属登录页面"""
    return render_template('customer_login.html', version=APP_VERSION)


@app.route('/api/customer_login', methods=['POST'])
def api_customer_login():
    """客户登录 API（仅允许 customer 角色）"""
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()

    client_ip = request.remote_addr or '0.0.0.0'
    if _is_login_rate_limited(client_ip):
        return jsonify({'success': False, 'message': f'登录尝试过于频繁，请 {_LOGIN_LOCKOUT_SECONDS // 60} 分钟后再试'}), 429

    if not username or not password:
        return jsonify({'success': False, 'message': '请输入用户名和密码'}), 400

    if verify_password(username, password):
        user = get_user(username)
        if user.role != 'customer':
            return jsonify({'success': False, 'message': '此入口仅限客户使用，员工请使用员工登录'}), 403
        if not user.is_active:
            return jsonify({'success': False, 'message': '账号待审核或已停用，请联系检测中心'}), 403
        login_user(user)
        log_action(username, 'login', '', '客户门户登录')
        _clear_login_attempts(client_ip)
        return jsonify({'success': True})

    _record_login_failure(client_ip)
    return jsonify({'success': False, 'message': '用户名或密码错误'}), 401


@app.route('/register')
def register_page():
    """客户自助注册页面"""
    return render_template('register.html', version=APP_VERSION)


@app.route('/api/register', methods=['POST'])
def api_register():
    """客户自助注册 API：创建待审核客户账号"""
    data = request.get_json(silent=True) or {}
    company = (data.get('company') or '').strip()
    contact_name = (data.get('contact_name') or '').strip()
    phone = (data.get('phone') or '').strip()
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    address = (data.get('address') or '').strip()

    # 校验
    if not company or not contact_name or not phone or not username or not password:
        return jsonify({'success': False, 'message': '请填写所有必填字段'})
    if len(username) < 3 or len(username) > 20:
        return jsonify({'success': False, 'message': '用户名需要3-20个字符'})
    import re as _re
    if not _re.match(r'^[a-zA-Z0-9_]+$', username):
        return jsonify({'success': False, 'message': '用户名只能包含字母、数字、下划线'})
    if len(password) < 6:
        return jsonify({'success': False, 'message': '密码至少6位'})
    if not _re.match(r'^1[3-9]\d{9}$', phone):
        return jsonify({'success': False, 'message': '请输入正确的手机号'})

    from database import get_db
    from werkzeug.security import generate_password_hash
    try:
        with get_db() as conn:
            # 检查用户名是否已存在
            existing = conn.execute('SELECT user_id FROM users WHERE user_id = ?', [username]).fetchone()
            if existing:
                return jsonify({'success': False, 'message': '该用户名已被注册'})

            # 创建账号：is_active=0（待审核）
            conn.execute(
                """INSERT INTO users (user_id, display_name, password_hash, role, department, is_active, client_name, created_at)
                   VALUES (?, ?, ?, 'customer', ?, 0, ?, datetime('now', 'localtime'))""",
                [username, contact_name, generate_password_hash(password, method='pbkdf2:sha256'), company, company]
            )

            # 记录注册附加信息到 customer_registrations 表
            _ensure_registration_table(conn)
            conn.execute(
                """INSERT INTO customer_registrations (username, company, contact_name, phone, address, status, created_at)
                   VALUES (?, ?, ?, ?, ?, 'pending', datetime('now', 'localtime'))""",
                [username, company, contact_name, phone, address]
            )
            conn.commit()

        log_action(username, 'customer_register', '', f'客户自助注册：{company} / {contact_name} / {phone}')
        notify_new_registration(company, contact_name, username)
        return jsonify({'success': True, 'message': '注册成功，等待审核'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'注册失败：{str(e)}'})


def _ensure_registration_table(conn):
    """确保 customer_registrations 表存在"""
    conn.execute("""CREATE TABLE IF NOT EXISTS customer_registrations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL,
        company TEXT NOT NULL,
        contact_name TEXT NOT NULL,
        phone TEXT NOT NULL,
        address TEXT DEFAULT '',
        status TEXT DEFAULT 'pending',
        reviewed_by TEXT DEFAULT '',
        reviewed_at TEXT DEFAULT '',
        reject_reason TEXT DEFAULT '',
        created_at TEXT NOT NULL
    )""")


@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    username = current_user.id
    role = getattr(current_user, 'role', 'admin')
    log_action(username, 'logout', '', '登出')
    logout_user()
    if role == 'customer':
        return redirect('/customer/login')
    return redirect(url_for('login_page'))


@app.route('/api/save', methods=['POST'])
@login_required
def api_save_compat():
    data = request.get_json(silent=True) or {}
    project = data.get('project') if isinstance(data, dict) else None
    if not isinstance(project, dict):
        project = data if isinstance(data, dict) else {}
    draft_id = ''
    if isinstance(data, dict):
        draft_id = (data.get('draft_id') or data.get('record_id') or project.get('record_id') or '').strip()
    # 第二层隔离：非本人草稿拒绝写入
    if draft_id:
        existing = _x_draft_path(draft_id)
        if existing.exists():
            try:
                old = json.loads(existing.read_text(encoding='utf-8'))
                old_inspector = (old.get('project') or {}).get('operator') or (old.get('project') or {}).get('inspector') or ''
                if old_inspector and old_inspector != current_user.id:
                    return jsonify({'success': False, 'error': '这是其他检测员的草稿，不能覆盖保存'}), 403
            except Exception:
                pass
    draft_kind = str(data.get('_draft_kind') or data.get('draft_kind') or '').strip().lower() or 'manual'
    payload = {
        'draft_id': draft_id or f"X1DRAFT_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        'schema_version': APP_VERSION,
        'source': 'x1-compat-save',
        'saved_at': _x_now(),
        'draft_kind': draft_kind,
        'project': project,
    }
    payload['project']['record_id'] = payload['draft_id']
    target = _x_draft_path(payload['draft_id'])
    with open(target, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    # 操作日志
    project_name = (project.get('project_name') or project.get('name') or '')[:50]
    log_action(
        current_user.id if current_user.is_authenticated else 'unknown',
        '保存记录',
        payload['draft_id'],
        f'{draft_kind} | {project_name}'
    )
    return jsonify({
        'success': True,
        'ok': True,
        'compat': True,
        'draft_id': payload['draft_id'],
        'record_id': payload['draft_id'],
        'saved_at': payload['saved_at'],
        'path': str(target)
    })


@app.route('/')
def landing():
    """公开首页 - 选择员工/客户入口"""
    if current_user.is_authenticated:
        if current_user.role == 'customer':
            return redirect('/customer')
        return redirect('/home')
    return render_template('landing.html')

@app.route('/home')
@login_required
def index():
    if current_user.role == 'customer':
        return redirect('/customer')
    return render_template('record_index.html', version=APP_VERSION, current_user=current_user)


@app.route('/admin')
@login_required
@require_permission('admin.access')
def admin():
    return render_template('admin.html', version=APP_VERSION)


@app.route('/api/list')
@login_required
def api_list_compat():
    """旧客户端兼容列表接口，返回草稿和导出记录。"""
    records = []

    def _draft_has_visible_content(project: dict, data: dict) -> bool:
        if not isinstance(project, dict):
            return False
        project_name = str(project.get('project_name', '') or '').strip()
        client_name = str(project.get('client_name', '') or '').strip()
        # 默认隐藏自动化测试草稿，避免前台列表被大量测试记录淹没
        if project_name.startswith('全变体测试-'):
            return False
        if client_name == '边界测试单位':
            return False
        if project_name and ('_low_' in project_name or '_high_' in project_name or project_name.startswith(('or_', 'bsl_', 'elec_', 'gmp_', 'food_'))):
            return False
        rooms = project.get('rooms') or []
        strong_fields = [
            project_name,
            client_name,
            project.get('contact_info', ''),
            project.get('project_address', ''),
            project.get('inspection_area', ''),
            project.get('detection_type', ''),
            project.get('detection_type_name', ''),
            project.get('remarks', ''),
        ]
        if any(str(v).strip() for v in strong_fields if v is not None):
            return True
        # 仅有日期或 rooms 的弱内容草稿不展示，避免空白卡片
        return False

    def _is_valid_export_record(export_id: str, proj: dict) -> bool:
        if not export_id.startswith('X1EXPORT_'):
            return False
        suffix = export_id[len('X1EXPORT_'):]
        if len(suffix) != 14 or not suffix.isdigit():
            return False
        if not isinstance(proj, dict):
            return False
        return any(str(proj.get(k, '')).strip() for k in ['project_name', 'client_name', 'report_number', 'detection_date'])
    
    # 1. 读取所有草稿
    auto_cutoff = datetime.now().timestamp() - 7 * 24 * 60 * 60
    for draft_file in RECORDS_DIR.glob('*.json'):
        try:
            with open(draft_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                project = data.get('project', {})
                if not _draft_has_visible_content(project, data):
                    continue
                draft_kind = str(data.get('draft_kind') or '').strip().lower() or 'manual'
                saved_at = data.get('updated_at', '') or data.get('created_at', '') or data.get('saved_at', '')
                if draft_kind == 'auto':
                    try:
                        if datetime.fromisoformat(saved_at.replace('Z', '+00:00')).timestamp() < auto_cutoff:
                            continue
                    except Exception:
                        pass
                records.append({
                    'record_id': data.get('draft_id', draft_file.stem),
                    'project_name': project.get('project_name', ''),
                    'client_name': project.get('client_name', ''),
                    'report_number': project.get('report_number', ''),
                    'detection_date': project.get('detection_date', ''),
                    'detection_type': project.get('detection_type', ''),
                    'detection_type_name': project.get('detection_type_name', ''),
                    'rooms': project.get('rooms', []) if isinstance(project.get('rooms', []), list) else [],
                    'room_count': len(project.get('rooms', []) if isinstance(project.get('rooms', []), list) else []),
                    'voided': bool(data.get('voided')),
                    'voided_at': data.get('voided_at', ''),
                    'voided_by': data.get('voided_by', ''),
                    'void_reason': data.get('void_reason', ''),
                    'inspector': project.get('operator', '') or project.get('inspector', ''),
                    'domain_name': project.get('domain_name', '') or project.get('domain', ''),
                    'schema_version': project.get('schema_version', '') or data.get('schema_version', ''),
                    'record_version': project.get('record_version', '') or data.get('record_version', ''),
                    'trace_id': project.get('trace_id', '') or data.get('trace_id', ''),
                    'normalized_at': project.get('normalized_at', '') or data.get('normalized_at', ''),
                    'status': 'draft',
                    'draft_kind': draft_kind,
                    'save_time': saved_at,
                    'export_info': None,
                    'report_info': None
                })
        except:
            pass
    
    # 2. 读取标准正式导出记录（以标准 export json 为准）
    for export_file in sorted(REPORTS_DIR.glob('X1EXPORT_*.json'), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            export_id = export_file.stem
            with open(export_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                ep = data.get('export_payload', data)
                proj = ep.get('project', {}) or {}
                room_from_export = ep.get('room', {}) if isinstance(ep.get('room', {}), dict) else {}
                rooms_from_project = proj.get('rooms', []) if isinstance(proj.get('rooms', []), list) else []
                normalized_rooms = rooms_from_project if rooms_from_project else ([room_from_export] if room_from_export else [])
                detection_type = proj.get('detection_type', '') or ep.get('export_type', '') or room_from_export.get('type_id', '')
                detection_type_name = proj.get('detection_type_name', '') or room_from_export.get('type_name', '')
                if not _is_valid_export_record(export_id, proj):
                    continue

                main_xlsx = REPORTS_DIR / f'{export_id}.xlsx'
                main_docx = REPORTS_DIR / f'{export_id}.docx'
                filled_docx = REPORTS_DIR / f'{export_id}.filled.docx'
                bound_docx = REPORTS_DIR / f'{export_id}.bound.docx'
                feishu = data.get('feishu', {}) or {}

                export_info = None
                report_info = None
                export_feishu_url = str((feishu.get('export', {}) or {}).get('feishu_url', '') or '').strip()
                report_feishu_url = str((feishu.get('report', {}) or {}).get('feishu_url', '') or '').strip()
                if export_feishu_url:
                    export_info = {
                        'feishu_url': export_feishu_url
                    }
                elif main_xlsx.exists():
                    export_info = {
                        'filename': main_xlsx.name,
                        'path': str(main_xlsx),
                    }
                elif main_docx.exists():
                    export_info = {
                        'filename': main_docx.name,
                        'path': str(main_docx),
                    }

                report_filename = filled_docx if filled_docx.exists() else (bound_docx if bound_docx.exists() else None)
                if report_feishu_url:
                    report_info = {
                        'feishu_url': report_feishu_url
                    }
                elif report_filename:
                    report_info = {
                        'filename': report_filename.name,
                        'path': str(report_filename),
                    }

                overall_status = data.get('overall_status')
                report_success = data.get('report_success')
                raw_record_success = data.get('raw_record_success')
                report_status = data.get('report_status')
                raw_record_status = data.get('raw_record_status')
                template_ready_value = data.get('template_ready', None)

                if report_success is None:
                    report_success = bool(report_info)
                if raw_record_success is None:
                    raw_record_success = bool(export_info)
                if report_status is None:
                    report_status = 'success' if report_success else ('blocked_template_missing' if template_ready_value is False and raw_record_success else 'missing')
                if raw_record_status is None:
                    raw_record_status = 'success' if raw_record_success else 'missing'
                if overall_status is None:
                    if report_success and raw_record_success:
                        overall_status = 'success'
                    elif raw_record_success and not report_success:
                        overall_status = 'partial_success'
                    else:
                        overall_status = 'failed'

                records.append({
                    'record_id': export_id,
                    'project_name': proj.get('project_name', ''),
                    'client_name': proj.get('client_name', ''),
                    'report_number': proj.get('report_number', ''),
                    'detection_date': proj.get('detection_date', ''),
                    'detection_type': detection_type,
                    'detection_type_name': detection_type_name,
                    'rooms': normalized_rooms,
                    'room_count': len(normalized_rooms),
                    'voided': bool(data.get('voided')),
                    'voided_at': data.get('voided_at', ''),
                    'voided_by': data.get('voided_by', ''),
                    'void_reason': data.get('void_reason', ''),
                    'inspector': proj.get('operator', '') or proj.get('inspector', ''),
                    'domain_name': proj.get('domain_name', '') or proj.get('domain', ''),
                    'schema_version': proj.get('schema_version', '') or data.get('schema_version', ''),
                    'record_version': proj.get('record_version', '') or data.get('record_version', ''),
                    'trace_id': proj.get('trace_id', '') or data.get('trace_id', ''),
                    'normalized_at': proj.get('normalized_at', '') or data.get('normalized_at', ''),
                    'status': 'completed',
                    'overall_status': overall_status,
                    'report_success': report_success,
                    'raw_record_success': raw_record_success,
                    'report_status': report_status,
                    'raw_record_status': raw_record_status,
                    'template_ready': template_ready_value,
                    'save_time': data.get('saved_at', '') or proj.get('saved_at', ''),
                    'export_info': export_info,
                    'report_info': report_info,
                })
        except:
            pass

    records = [r for r in records if can_view_record(current_user, {'inspector_name': r.get('inspector', '')})]
    # 统一规则：暂存记录与报告记录一律按 save_time 倒序返回（最新在最前）
    records.sort(key=lambda r: r.get('save_time') or '', reverse=True)
    
    return jsonify({
        'success': True,
        'compat': True,
        'compat_route': '/api/list',
        'exports': [
            {
                'name': r['export_info']['filename'],
                'path': r['export_info']['path'],
                'suffix': Path(r['export_info']['filename']).suffix.lower() if r['export_info'].get('filename') else ''
            }
            for r in records if r.get('export_info') and r['export_info'].get('filename')
        ],
        'records': records,
    })




@app.route('/api/x/passbox-sample')
def api_x_passbox_sample():
    return jsonify({
        'success': True,
        'sample': {
            'room_id': 'r1',
            'type_id': 'pass_box',
            'type_name': '传递窗',
            'domain': 'pharma',
            'room_name': 'X1传递窗样板',
            'level_name': '无等级要求',
            'clean_class': '无等级要求',
            'basis': ['GB 50591-2010'],
            'judgement': ['JG/T 382-2012', 'GB 50591-2010'],
            'params': [],
            'summary': {
                'result_state': '',
                'judgement_active': ['JG/T 382-2012'],
                'basis_primary': 'GB 50591-2010',
                'judgement_primary': 'JG/T 382-2012'
            },
            'context': {}
        }
    })

@app.route('/download/<filename>')
@login_required
def download_file(filename):
    """下载导出的报告文件"""
    if not _setting_enabled('security.allow_file_download', True):
        return jsonify({'success': False, 'error': '系统设置已禁止文件下载'}), 403
    import re
    # 安全检查：只允许合法文件名
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'success': False, 'error': '非法文件名'}), 400
    if not _can_access_file_by_name(filename):
        return jsonify({'success': False, 'error': '无权访问该文件'}), 403
    
    file_path = REPORTS_DIR / filename
    if not file_path.exists():
        # 也在 records 目录中找
        file_path = RECORDS_DIR / filename
    if not file_path.exists():
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    
    return send_file(str(file_path), as_attachment=True, download_name=filename)


@app.route('/admin/api/open_file/<filename>', methods=['POST'])
@login_required
@require_permission('admin.records.open_local')
def admin_api_open_file(filename):
    """在服务器本机优先用 Word/Excel 打开文件"""
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'success': False, 'error': '非法文件名'}), 400
    if not _can_access_file_by_name(filename):
        return jsonify({'success': False, 'error': '无权访问该文件'}), 403

    file_path = REPORTS_DIR / filename
    if not file_path.exists():
        file_path = RECORDS_DIR / filename
    if not file_path.exists():
        return jsonify({'success': False, 'error': '文件不存在'}), 404

    if not _is_desktop_mode():
        return jsonify({'success': False, 'error': '当前为 server 模式，已禁用本机打开文件；请改用下载方式获取文件'}), 409

    try:
        import subprocess
        suffix = file_path.suffix.lower()
        app_name = 'WPS Office'
        if suffix in ('.docx', '.doc', '.xlsx', '.xls', '.csv', '.pptx', '.ppt'):
            subprocess.Popen(['open', '-a', 'wpsoffice', str(file_path)])
        else:
            app_name = '系统默认软件'
            subprocess.Popen(['open', str(file_path)])
        return jsonify({'success': True, 'message': f'已调用 {app_name} 打开文件', 'app_name': app_name, 'filename': filename, 'path': str(file_path)})
    except Exception as e:
        return jsonify({'success': False, 'error': f'打开文件失败: {e}'}), 500


@app.route('/admin/api/download_feishu_file')
@login_required
@require_permission('admin.records.open_feishu')
def admin_api_download_feishu_file():
    """从飞书云盘下载文件并返回给当前浏览器（多人多端场景）
    
    action=open 时使用 inline 模式，让手机系统弹出“用什么 App 打开”选择器。
    默认 attachment 模式，直接下载到设备。
    """
    file_token = (request.args.get('file_token') or '').strip()
    if not file_token:
        return jsonify({'success': False, 'error': '缺少 file_token'}), 400
    result = download_file_content_from_feishu(file_token)
    if not result.get('success'):
        return jsonify({'success': False, 'error': result.get('error', '飞书下载失败')}), 500
    filename = result['filename']
    content = result['content']
    content_type = result.get('content_type') or 'application/octet-stream'
    # 根据扩展名设置正确的 MIME type，因为飞书可能返回 octet-stream
    _suffix = Path(filename).suffix.lower()
    _mime_map = {
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.doc': 'application/msword',
        '.xls': 'application/vnd.ms-excel',
        '.pdf': 'application/pdf',
    }
    content_type = _mime_map.get(_suffix, content_type)
    action = (request.args.get('action') or '').strip().lower()
    as_attachment = (action != 'open')  # open = inline, 其他 = attachment
    from io import BytesIO
    return send_file(
        BytesIO(content),
        mimetype=content_type,
        as_attachment=as_attachment,
        download_name=filename,
        max_age=0
    )


@app.route('/admin/api/open_feishu_file', methods=['POST'])
@login_required
@require_permission('admin.records.open_feishu')
def admin_api_open_feishu_file():
    """从飞书云盘下载文件并用 WPS 打开"""
    data = request.get_json(silent=True) or {}
    file_token = data.get('file_token', '').strip()
    if not file_token:
        return jsonify({'success': False, 'error': '缺少 file_token'}), 400
    result = download_file_from_feishu(file_token)
    if not result.get('success'):
        return jsonify({'success': False, 'error': result.get('error', '飞书下载失败')}), 500
    file_path = result['path']
    if not _is_desktop_mode():
        return jsonify({'success': False, 'error': '当前为 server 模式，已禁用本机打开飞书文件；请先下载后手动打开'}), 409
    try:
        import subprocess
        subprocess.Popen(['open', '-a', 'wpsoffice', file_path])
        return jsonify({'success': True, 'message': f'已从飞书下载并用 WPS 打开: {result["filename"]}', 'filename': result['filename'], 'size': result['size']})
    except Exception as e:
        return jsonify({'success': False, 'error': f'打开失败: {e}'}), 500


@app.route('/api/preview/<filename>')
@login_required
@require_permission('files.preview.own')
def preview_file(filename):
    """在线预览 docx/xlsx 文件（返回 HTML）"""
    if not _setting_enabled('security.allow_file_preview', True):
        return jsonify({'success': False, 'error': '系统设置已禁止文件预览'}), 403
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'success': False, 'error': '非法文件名'}), 400
    if not _can_access_file_by_name(filename):
        return jsonify({'success': False, 'error': '无权访问该文件'}), 403
    
    file_path = REPORTS_DIR / filename
    if not file_path.exists():
        file_path = RECORDS_DIR / filename
    if not file_path.exists():
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    
    suffix = file_path.suffix.lower()
    try:
        if suffix == '.docx':
            import mammoth
            with open(str(file_path), 'rb') as f:
                result = mammoth.convert_to_html(f)
                html_content = result.value
            return jsonify({
                'success': True,
                'html': html_content,
                'filename': filename,
                'file_size': file_path.stat().st_size,
                'type': 'docx'
            })
        elif suffix in ('.xlsx', '.xls'):
            import html
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            html_parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                html_parts.append(f'<h3 style="margin:16px 0 8px;color:#1890ff">{html.escape(str(sheet_name))}</h3>')
                html_parts.append('<table style="border-collapse:collapse;width:100%;margin-bottom:16px">')
                for row in ws.iter_rows(values_only=False):
                    html_parts.append('<tr>')
                    for cell in row:
                        val = cell.value if cell.value is not None else ''
                        safe_val = html.escape(str(val))
                        # 处理合并单元格
                        style = 'border:1px solid #d9d9d9;padding:6px 8px;font-size:13px'
                        if isinstance(val, (int, float)):
                            style += ';text-align:right'
                        html_parts.append(f'<td style="{style}">{safe_val}</td>')
                    html_parts.append('</tr>')
                html_parts.append('</table>')
            wb.close()
            return jsonify({
                'success': True,
                'html': ''.join(html_parts),
                'filename': filename,
                'file_size': file_path.stat().st_size,
                'type': 'xlsx'
            })
        else:
            return jsonify({'success': False, 'error': f'不支持预览 {suffix} 格式'}), 400
    except ImportError as e:
        return jsonify({'success': False, 'error': f'缺少依赖库: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': f'预览失败: {str(e)}'}), 500




# ==================== Blueprint 注册 ====================
# ==================== /admin/api/backups 备份列表 ====================
@app.route('/admin/api/backups')
@login_required
@require_permission('admin.monitor.view')
def admin_api_backups():
    """返回备份目录中的所有备份文件列表"""
    import os
    from helpers.settings_utils import _load_system_settings
    settings_values = _load_system_settings()
    backup_dir = str(Path(str(settings_values.get('paths.backup_dir', {}).get('value', BASE_DIR / 'backups'))).expanduser())
    items = []
    try:
        for f in sorted(os.listdir(backup_dir), reverse=True):
            if f.endswith('.tar.gz') or f.endswith('.zip'):
                fpath = os.path.join(backup_dir, f)
                stat = os.stat(fpath)
                size_mb = round(stat.st_size / 1024 / 1024, 1)
                mtime = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                btype = '自动' if '_auto_' in f else '手动'
                items.append({'filename': f, 'size_mb': size_mb, 'time': mtime, 'type': btype})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    return jsonify({'success': True, 'items': items, 'backup_dir': backup_dir})

# ==================== /api/x/health 健康检查 ====================
@app.route('/api/x/health')
def api_x_health():
    remote = (request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
    if remote not in {'127.0.0.1', '::1', 'localhost'} and not current_user.is_authenticated:
        return redirect(url_for('login_page', next=request.url))
    try:
        cpu_percent = psutil.cpu_percent(interval=0.1)
        memory = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        system_info = {
            'cpu_percent': round(cpu_percent, 1),
            'memory_percent': round(memory.percent, 1),
            'memory_used_gb': round(memory.used / (1024**3), 2),
            'memory_total_gb': round(memory.total / (1024**3), 2),
            'disk_percent': round(disk.percent, 1),
            'disk_used_gb': round(disk.used / (1024**3), 2),
            'disk_total_gb': round(disk.total / (1024**3), 2),
        }
    except:
        system_info = {}
    from helpers.settings_utils import _load_system_settings, _get_latest_backup
    settings_values = _load_system_settings()
    backup_dir_path = Path(str(settings_values.get('paths.backup_dir', {}).get('value', BASE_DIR / 'backups'))).expanduser()
    return jsonify({
        'success': True,
        'app': 'X1',
        'version': APP_VERSION,
        'host': APP_HOST,
        'port': APP_PORT,
        'base_dir': str(BASE_DIR),
        'template_base': str(TEMPLATE_BASE),
        'records_dir': str(RECORDS_DIR),
        'reports_dir': str(REPORTS_DIR),
        'logs_dir': str(LOGS_DIR),
        'cache_dir': str(CACHE_DIR),
        'temp_dir': str(TEMP_DIR),
        'uploads_dir': str(UPLOADS_DIR),
        'template_config': str(TEMPLATE_CONFIG_FILE),
        'backup_dir': str(backup_dir_path),
        'backup_latest': _get_latest_backup(backup_dir_path),
        'system': system_info,
        'isolation': {
            'shares_v_runtime': False,
            'shares_t_runtime': False,
            'api_prefix': '/api/x/*',
        },
    })


# ==================== 后台页面路由 ====================
@app.route('/admin/standards')
@login_required
def admin_standards():
    return render_template('standards.html')


@app.route('/admin/monitor')
@login_required
def admin_monitor():
    return render_template('monitor.html')


@app.route('/admin/api/docs/<doc_name>')
@login_required
def admin_api_docs(doc_name):
    allowed = {'ARCHITECTURE': 'ARCHITECTURE.md', 'API': 'API.md'}
    if doc_name not in allowed:
        return jsonify({'error': '文档不存在'}), 404
    doc_path = Path(__file__).parent / 'docs' / allowed[doc_name]
    if not doc_path.exists():
        return jsonify({'error': '文档文件不存在'}), 404
    return jsonify({'content': doc_path.read_text(encoding='utf-8')})


@app.route('/admin/api/workspace_doc')
@login_required
def admin_api_workspace_doc():
    allowed = {
        'X1_系统当前版本说明.md', 'X1_版本号管理规则.md',
        'ARCHITECTURE.md', 'API.md', 'HOST_MODE.md',
        'X1 常见问题排障手册.md', 'X1 运维启动-停止-验活说明.md',
        'X1 部署与迁移说明.md', 'X1 飞书上传失败治理 SOP.md',
        'X1_全量代码审计报告_2026-05-16_v2.md',
        'X1_架构重构计划_2026-05-16.md',
        'X1_飞书月目录自动切换运维说明_2026-05-12.md',
    }
    requested = request.args.get('path', '').strip()
    if not requested:
        return jsonify({'error': '缺少 path 参数'}), 400
    doc_path = Path(requested)
    if doc_path.name not in allowed:
        return jsonify({'error': '文档未授权'}), 403
    if not doc_path.exists() or not doc_path.is_file():
        return jsonify({'error': '文档不存在'}), 404
    return jsonify({'content': doc_path.read_text(encoding='utf-8')})


@app.route('/admin/api/records/<record_id>', methods=['DELETE'])
@login_required
def admin_api_delete_record(record_id):
    if not _setting_enabled('security.allow_delete_record', True):
        return jsonify({'success': False, 'error': '系统设置已禁止删除记录'}), 403
    ok, msg = _soft_delete_record(record_id)
    if not ok:
        return jsonify({'success': False, 'error': msg}), 404
    log_action(current_user.id if current_user.is_authenticated else 'unknown', '删除记录', record_id, msg)
    return jsonify({'success': True, 'message': msg})


@app.route('/admin/api/cleanup_trash', methods=['POST'])
@login_required
def admin_api_cleanup_trash():
    if not _setting_enabled('security.allow_cleanup_trash', True):
        return jsonify({'success': False, 'error': '系统设置已禁止清空回收站'}), 403
    data = request.get_json(silent=True) or {}
    days = int(data.get('days', 30))
    result = cleanup_trash(days)
    log_action(current_user.id if current_user.is_authenticated else 'unknown', '清理回收站', '', f"清理 {result['deleted_count']} 个文件")
    return jsonify({'success': True, **result})


# ==================== Blueprint 注册 ====================
from routes import register_blueprints
register_blueprints(app)

if __name__ == '__main__':
    print(f"🚀 X1 skeleton running at http://{APP_HOST}:{APP_PORT}")
    app.run(host=APP_HOST, port=APP_PORT, debug=False)


